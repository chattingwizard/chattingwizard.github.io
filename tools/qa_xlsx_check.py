"""QA Check for ALL *_Complete_Infloww.xlsx files in CW-ScriptManager."""
import sys
import os

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

BASE = r"c:\Users\34683\CW-ScriptManager"
PATTERN = "_Complete_Infloww.xlsx"
MAX_NAME = 64
MAX_TEXT = 1000
MAX_NOTE = 246

# â”€â”€ Find files â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
xlsx_files = []
for root, dirs, files in os.walk(BASE):
    dirs[:] = [d for d in dirs if d not in [".git", "node_modules", "__pycache__", ".cursor"]]
    for fname in files:
        if fname.endswith(PATTERN):
            xlsx_files.append(os.path.join(root, fname))
xlsx_files.sort()

total_files = len(xlsx_files)
total_sheets = 0
total_scripts = 0
total_violations = 0
files_with_issues = 0

SEP = "=" * 80
THIN = "-" * 60

print(SEP)
print("  QA CHECK - Infloww XLSX Files")
print(f"  Base: {BASE}")
print(f"  Files found: {total_files}")
print(SEP)

if total_files == 0:
    print("\nNo *_Complete_Infloww.xlsx files found.")
    sys.exit(0)

# â”€â”€ Process each file â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
for fpath in xlsx_files:
    rel = os.path.relpath(fpath, BASE)
    file_violations = []

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        print(f"\n[ERROR] Cannot open: {rel} - {e}")
        total_violations += 1
        files_with_issues += 1
        continue

    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)
    total_sheets += num_sheets

    print(f"\n{THIN}")
    print(f"FILE: {rel}")
    print(f"  Sheets ({num_sheets}): {', '.join(sheet_names)}")

    file_script_count = 0

    for ws in wb.worksheets:
        sheet_title = ws.title
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            file_violations.append(f"  [{sheet_title}] EMPTY SHEET - no rows at all")
            continue

        # â”€â”€ Header check â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        header = rows[0]
        header_clean = [str(h).strip() if h else "" for h in header[:3]]
        expected = ["Name", "Text", "Note"]
        has_good_header = (header_clean == expected)
        if not has_good_header:
            file_violations.append(
                f"  [{sheet_title}] BAD HEADERS - expected {expected}, got {header_clean}"
            )

        data_rows = rows[1:]  # skip header row
        row_count = 0
        names_seen = {}

        for idx, row in enumerate(data_rows, start=2):
            # Pad to at least 3 columns
            padded = list(row) + [None] * max(0, 3 - len(row))
            name_val = padded[0]
            text_val = padded[1]
            note_val = padded[2]

            # Skip completely empty rows
            if not any(padded[:3]):
                continue

            row_count += 1
            name_str = str(name_val).strip() if name_val else ""
            text_str = str(text_val).strip() if text_val else ""
            note_str = str(note_val).strip() if note_val else ""

            # Length checks
            if name_str and len(name_str) > MAX_NAME:
                file_violations.append(
                    f"  [{sheet_title}] Row {idx}: Name exceeds {MAX_NAME} chars "
                    f"({len(name_str)}): \"{name_str[:70]}...\""
                )
            if text_str and len(text_str) > MAX_TEXT:
                file_violations.append(
                    f"  [{sheet_title}] Row {idx}: Text exceeds {MAX_TEXT} chars "
                    f"({len(text_str)}): \"{text_str[:80]}...\""
                )
            if note_str and len(note_str) > MAX_NOTE:
                file_violations.append(
                    f"  [{sheet_title}] Row {idx}: Note exceeds {MAX_NOTE} chars "
                    f"({len(note_str)}): \"{note_str[:80]}...\""
                )

            # Empty Text with a Name present
            if name_str and not text_str:
                file_violations.append(
                    f"  [{sheet_title}] Row {idx}: Empty Text for Name: \"{name_str[:50]}\""
                )

            # Duplicate Names within same sheet
            if name_str:
                if name_str in names_seen:
                    file_violations.append(
                        f"  [{sheet_title}] Row {idx}: Duplicate Name: \"{name_str[:50]}\" "
                        f"(first at row {names_seen[name_str]})"
                    )
                else:
                    names_seen[name_str] = idx

        file_script_count += row_count
        print(f"  [{sheet_title}] {row_count} scripts | Cols: {header_clean}")

    wb.close()
    total_scripts += file_script_count
    print(f"  Total scripts in file: {file_script_count}")

    if file_violations:
        files_with_issues += 1
        total_violations += len(file_violations)
        print(f"\n  !! {len(file_violations)} VIOLATION(S):")
        for v in file_violations:
            print(f"    {v}")
    else:
        print("  [OK] No issues found")

# â”€â”€ Summary â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
print(f"\n{SEP}")
print("  SUMMARY")
print(SEP)
print(f"  Total XLSX files:   {total_files}")
print(f"  Total sheets:       {total_sheets}")
print(f"  Total scripts:      {total_scripts}")
print(f"  Total violations:   {total_violations}")
print(f"  Files with issues:  {files_with_issues} / {total_files}")
print()
if total_violations == 0:
    print("  ALL CLEAN - no violations found.")
else:
    print(f"  !! {total_violations} VIOLATION(S) FOUND - review above for details.")
print(SEP)
