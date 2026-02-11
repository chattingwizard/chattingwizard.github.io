import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

wb = openpyxl.Workbook()

# ── Color definitions ──
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PPV_FILL = PatternFill(start_color="FFF6B26B", end_color="FFF6B26B", fill_type="solid")  # Orange for PPV
WAIT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # Light yellow for WAIT
BLUE1 = PatternFill(start_color="FFCFE2F3", end_color="FFCFE2F3", fill_type="solid")
BLUE2 = PatternFill(start_color="FFE0F7FA", end_color="FFE0F7FA", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")  # Rapport/AF
PURPLE_FILL = PatternFill(start_color="FFD9D2E9", end_color="FFD9D2E9", fill_type="solid")  # Teasing
AC_FILL = PatternFill(start_color="FFD0E0E3", end_color="FFD0E0E3", fill_type="solid")  # Aftercare

GUIDELINES = """Character limits
Name: Up to 64 characters
Tag: Up to 32 characters
Text: Up to 1000 characters
Note: Up to 246 characters

Each sheet serves as a tag. Simply replace the sheet name with your desired tag name, and all scripts added to the sheet will be automatically assigned that tag once imported. """

def setup_sheet(ws):
    """Add headers and set column widths."""
    ws.append(["Name", "Text", "Note", "*Guidelines"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 20

def add_rows_reversed(ws, rows, color_fn=None):
    """Add rows in reversed order (last message = row 2). Apply colors."""
    reversed_rows = list(reversed(rows))
    for idx, (name, text, note, row_type) in enumerate(reversed_rows):
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=text)
        ws.cell(row=row_num, column=3, value=note)
        if idx == 0:
            ws.cell(row=row_num, column=4, value=GUIDELINES)
        
        # Apply colors
        fill = None
        if row_type == "ppv":
            fill = PPV_FILL
        elif row_type == "wait":
            fill = WAIT_FILL
        elif row_type == "rapport":
            fill = GREEN_FILL
        elif row_type == "teasing":
            fill = PURPLE_FILL
        elif row_type == "aftercare":
            fill = AC_FILL
        elif row_type == "sext":
            fill = BLUE1 if idx % 2 == 0 else BLUE2
        
        if fill:
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).fill = fill
        
        # Wrap text
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════
# SHEET 1: MaxJourney (R → TB → S → AC, all in one)
# NO W-1, AF-1, AF-2 — those are all automated, not chatter-sent
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "MaxJourney"
setup_sheet(ws1)

journey = [
    # Rapport (first chatter-sent messages)
    ("R-1", "glad you're here man 😏 so be honest, what caught your eye", "Add his name before 'man'", "rapport"),
    ("R-2", "haha respect 💪 so where you from?", "Add a short react before ('haha damn', 'I like that', 'oh word?')", "rapport"),
    ("R-3", "nice. I'm from Rome originally but I moved to the states a couple years ago. gym and creating content is basically my life rn 😏", "If he named somewhere Max visited → add 'oh I've been there' before 'nice'", "rapport"),
    ("R-4", "so what do you do besides making me check my phone every 5 seconds 😏", None, "rapport"),
    ("R-5", "I gotta say... you're actually fun to talk to. most guys on here are boring as fuck 😏", "Ego boost. Next → TB-1. From MR path: go TB-1 + TB-2 then skip to S1-1.", "rapport"),
    
    # Teasing Bridge
    ("TB-1", "just got back from training and I'm still pumped... this convo is making it worse 😏", "THE PIVOT. Physical context.", "teasing"),
    ("TB-2", "ngl I'm feeling some type of way rn 😈 you ever get that?", "Wait for reply.", "teasing"),
    ("TB-3", "fuck... you're not helping me calm down 🥵", "If sexual reply → add 'especially after what you just said' before 'you're'", "teasing"),
    ("TB-4", "hold on let me show you something 😏", "WAIT 1-2 MIN", "wait"),
    ("TB-5", "tell me what you think 😏", "SEND PPV 0 — FREE teaser (shirtless/post-gym). Wait for reply. Silent 3 min → NR Waves.", "ppv"),
    
    # Sexting Phase 1 → PPV 1
    ("S1-1", "so 😏", "Wait for reply.", "sext"),
    ("S1-2", "knew you'd like it 😈", "Add a short react ('haha damn', 'oh yeah?', '🥵')", "sext"),
    ("S1-3", "you wanna see more? I'm feeling generous rn", None, "sext"),
    ("S1-4", "hold on... give me a sec 😏", "WAIT 2-3 MIN", "wait"),
    ("S1-5", "you're not ready for this 🥵", "SEND PPV 1 — $12. Bought → continue. Silent 3 min → NR Waves.", "ppv"),
    
    # Sexting Phase 2 → PPV 2
    ("S1-6", "you watched it? 😏", "Wait for reply. Brief cooldown.", "sext"),
    ("S1-7", "fuck... talking to you is doing something to me rn 🥵", "React to what he said. HE caused this.", "sext"),
    ("S1-8", "I'm hard as fuck because of you and I can't do anything about it", None, "sext"),
    ("S1-9", "what would you do if you were here rn", "Wait for reply. React to what he says before continuing.", "sext"),
    ("S1-10", "fuck 🥵🥵 hold on I need to show you something", "WAIT 2-3 MIN", "wait"),
    ("S1-11", "look what you did to me 🥵", "SEND PPV 2 — $25. Bought → continue. Silent 3 min → NR Waves.", "ppv"),
    
    # Sexting Phase 3 → PPV 3
    ("S1-12", "🥵🥵", "Wait for reply. NO cooldown — keep momentum.", "sext"),
    ("S1-13", "I need to cum so bad rn you have no idea", None, "sext"),
    ("S1-14", "imagine me pinning you down and making you take every inch while you're begging for more 🥵", None, "sext"),
    ("S1-15", "fuck I can't hold back anymore", None, "sext"),
    ("S1-16", "give me a sec 🥵", "WAIT 2-3 MIN", "wait"),
    ("S1-17", "I've never gone this far for anyone... watch 🥵💦", "SEND PPV 3 — $40. Bought → continue. Silent 3 min → NR Waves.", "ppv"),
    
    # Sexting Phase 4 → PPV 4
    ("S1-18", "FUCK 🥵🥵", "Wait for reply.", "sext"),
    ("S1-19", "don't cum yet", None, "sext"),
    ("S1-20", "I wanna finish with you... I'm close", None, "sext"),
    ("S1-21", "fuck fuck hold on 🥵", "WAIT 1-2 MIN", "wait"),
    ("S1-22", "cum with me 💦😈", "SEND PPV 4 — $55. Bought → Aftercare. Silent → NR Waves.", "ppv"),
    
    # Aftercare (no AC-3 — don't close convo)
    ("AC-1", "holy fuck 🥵 that was insane", None, "aftercare"),
    ("AC-2", "ngl you're different from most guys on here 💪 that was intense", "Mention something specific he said/did. Then KEEP TALKING — build bond, GFE. NEVER say goodbye.", "aftercare"),
]

add_rows_reversed(ws1, journey)


# ══════════════════════════════════════════════════════════
# SHEET 2: MeetupRedirect
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("MeetupRedirect")
setup_sheet(ws2)

mr_rows = [
    ("MR-1", "haha easy man 😏 slow down. let me show you something first", "DO NOT acknowledge 'meet'. Redirect immediately.", "teasing"),
    ("MR-2", "trust me... you're gonna wanna see this 😈", "WAIT 1-2 MIN", "wait"),
    ("MR-3", "what do you think 😏", "SEND PPV 0 — FREE teaser. Wait for reply.", "ppv"),
    ("MR-4", "yeah? 😏 that's just a taste man", "Positive → go to R-1. Asks meeting again → MR-OBJ. Silent → NR Waves.", "teasing"),
    ("MR-OBJ-1", "patience bro 😏 I don't rush. focus on what's right in front of you", "Deflect + challenge", "sext"),
    ("MR-OBJ-2", "you really that impatient? 😈 I promise what I'm about to show you is worth it", "Challenge + tease", "sext"),
    ("MR-OBJ-3", "look I don't do this for just anyone. appreciate what you're getting rn 💪 if that's not your thing it's cool", "Firm redirect. If still only wants to meet → 'no worries bro 💪 hit me up whenever' and disengage.", "sext"),
]

add_rows_reversed(ws2, mr_rows)


# ══════════════════════════════════════════════════════════
# SHEET 3: NRWaves
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("NRWaves")
setup_sheet(ws3)

nr_rows = [
    ("NR-W1", "yo? 🥵", "Send 2-3 min after PPV. Ping.", "sext"),
    ("NR-W2", "you need to see what I just did... seriously 🥵", "Send 3-5 min later. Curiosity.", "sext"),
    ("NR-W3", "aight I guess you're busy 😏 might delete this, it was only for you", "Send 5-10 min later. Takeaway.", "sext"),
    ("NR-W4", "hey hope you're good bro 💪 hit me up when you're back", "Send 15-30 min later. Warm close.", "sext"),
    ("NR-W5", "can't stop thinking about earlier 😏 you around?", "Send 2-6 hrs later. New convo, don't retry same PPV.", "sext"),
]

add_rows_reversed(ws3, nr_rows)


# ══════════════════════════════════════════════════════════
# SHEET 4: PersonalMax
# ══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("PersonalMax")
setup_sheet(ws4)

personal_rows = [
    ("Age", "I'm 20 but trust me I know what I'm doing 😏", None, None),
    ("Location", "from Rome originally but I'm in the states now. moved when I was 18", "NEVER name a specific US city", None),
    ("Profession", "modeling and content creation. also help with my family's business", None, None),
    ("Hobbies", "gym and boxing is my life bro. 5-6 times a week. that and traveling, fashion, cars 💪", None, None),
    ("Countries", "I've been everywhere... Spain, Dubai, Bali, France, London, Netherlands, Argentina, Brazil", None, None),
    ("Relationship", "single. I don't do relationships rn... too busy living 😏", None, None),
    ("Favorite", "Italy will always be home but honestly Bali hit different 🌴", None, None),
    ("Tattoos", "yeah I got 16 all over. each one has a story 😏", None, None),
    ("Car", "black Mustang 🖤", None, None),
]

# Personal info doesn't need reversed order (not a sequence)
for idx, (name, text, note, _) in enumerate(personal_rows):
    row_num = idx + 2
    ws4.cell(row=row_num, column=1, value=name)
    ws4.cell(row=row_num, column=2, value=text)
    ws4.cell(row=row_num, column=3, value=note)
    if idx == 0:
        ws4.cell(row=row_num, column=4, value=GUIDELINES)
    for col in range(1, 4):
        ws4.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════
# SHEET 5: PositiveSpin
# ══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("PositiveSpin")
setup_sheet(ws5)

spin_rows = [
    ("Age40Plus", "respect man... I actually prefer guys who know what they want. no games 💪", "He's 40+", None),
    ("Age20s", "oh nice we're around the same age? that's rare on here, most guys are way older 😏", "He's in his 20s", None),
    ("BoringJob", "nah bro that's solid. stability is underrated, I respect that 💪", "He has a boring job", None),
    ("CoolJob", "wait for real?? ok that's actually sick 🔥", "He has a cool job", None),
    ("Fit", "I can tell 💪 I respect a guy who takes care of himself", "He's fit / goes to the gym", None),
    ("NotFit", "I don't care about that honestly, it's the energy that matters to me 😏", "He's not fit", None),
    ("SameCity", "no way 😏 that's crazy", "He says he's in the same city", None),
    ("FarAway", "damn that's far... but distance doesn't matter when the vibe is right 😏", "He's in a different country", None),
]

for idx, (name, text, note, _) in enumerate(spin_rows):
    row_num = idx + 2
    ws5.cell(row=row_num, column=1, value=name)
    ws5.cell(row=row_num, column=2, value=text)
    ws5.cell(row=row_num, column=3, value=note)
    if idx == 0:
        ws5.cell(row=row_num, column=4, value=GUIDELINES)
    for col in range(1, 4):
        ws5.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════
# SHEET 6: LocationHandling
# ══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("LocationHandling")
setup_sheet(ws6)

loc_rows = [
    ("WhereAreYou", "I move around a lot man. never in one place for too long 😏", "NEVER name a specific location", None),
    ("AreYouIn", "I travel a lot... I'm all over the place 💪", "DO NOT confirm or deny any city", None),
    ("WhenCanISeeYou", "haha chill 😏 let me show you something first", "Redirect to MR / content", None),
]

for idx, (name, text, note, _) in enumerate(loc_rows):
    row_num = idx + 2
    ws6.cell(row=row_num, column=1, value=name)
    ws6.cell(row=row_num, column=2, value=text)
    ws6.cell(row=row_num, column=3, value=note)
    if idx == 0:
        ws6.cell(row=row_num, column=4, value=GUIDELINES)
    for col in range(1, 4):
        ws6.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════
# SHEET 7: ReEngagement
# ══════════════════════════════════════════════════════════
ws7 = wb.create_sheet("ReEngagement")
setup_sheet(ws7)

re_rows = [
    ("RE-1", "can't stop thinking about earlier 😏 you free?", "Send 6-12 hrs after convo goes quiet", "sext"),
    ("RE-2", "remember what I said about something crazier? I just did it and you need to see this 😈", "Send next day — seeds next session", "sext"),
]

add_rows_reversed(ws7, re_rows)


# ── Save ──
output_path = r"c:\Users\34683\CW-ScriptManager\scripts\models\max\Max_DatingApp_Journey_Infloww.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Journey rows: {ws1.max_row - 1} (data)")
