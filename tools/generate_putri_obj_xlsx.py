import sys, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()

# ── STYLES ──
HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(bold=True, color="e6edf3", size=11)
PRICE_FILL = PatternFill("solid", fgColor="2d1a1a")
DISCOUNT_FILL = PatternFill("solid", fgColor="2d2a1a")
FREE_FILL = PatternFill("solid", fgColor="1a2d1a")
NOMONEY_FILL = PatternFill("solid", fgColor="1a1a2d")
NOPPV_FILL = PatternFill("solid", fgColor="2d1a2d")
CARD_FILL = PatternFill("solid", fgColor="1a2d2d")
RES_FILL = PatternFill("solid", fgColor="1f1a2d")
SIT_FILL = PatternFill("solid", fgColor="1a2a1a")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="30363d")
)

def setup_sheet(ws):
    ws.append(["Name", "Text", "Note", "*Guidelines"])
    for col_letter in ["A","B","C","D"]:
        ws[f"{col_letter}1"].fill = HEADER_FILL
        ws[f"{col_letter}1"].font = HEADER_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 30

def add_rows(ws, rows):
    """Add rows in REVERSED order (Infloww convention). Each row: (name, text, note, fill)"""
    for name, text, note, fill in reversed(rows):
        ws.append([name, text, note, ""])
        row_num = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

# ═══════════════════════════════════════════════════════
# SHEET 1: ObjHandling
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "ObjHandling"
setup_sheet(ws1)

obj_data = [
    # ── OBJ-PRICE ──
    ("PRICE 1a Reframe", "babe that's less than a smoothie after my ride 😏 and trust me this hits wayyyy harder", "REFRAME. 1st try when he says too expensive. Still no → 1b/1c or Step 2.", PRICE_FILL),
    ("PRICE 1b Reframe", "baby it's like the price of your morning coffee ☕ except this one keeps you up all night 😏", "REFRAME variant. 1st try. Still no → Step 2 (FOMO).", PRICE_FILL),
    ("PRICE 1c Reframe", "that's literally less than an uber ride 🙈 and I promise you'll enjoy this trip way more", "REFRAME variant. 1st try. Still no → Step 2 (FOMO).", PRICE_FILL),
    ("PRICE 2a FOMO", "I'm literally only in this mood because of YOU rn 🥵 idk when I'll feel like this again... I never do this", "FOMO TEMPORAL. Step 2. He said no to reframe. Still no → Step 3 (Challenge).", PRICE_FILL),
    ("PRICE 2b FOMO", "this mood I'm in rn? it's not gonna last forever 🥵 and I really want you to be the one who sees it", "FOMO TEMPORAL variant. Step 2. Still no → Step 3 (Challenge).", PRICE_FILL),
    ("PRICE 3a Challenge", "hmm maybe you're not ready for what I did in this one 😏", "CHALLENGE. Step 3. Attacks ego. Still no → Step 4 (Downgrade).", PRICE_FILL),
    ("PRICE 3b Challenge", "honestly? most guys can't handle what I just recorded 😈 I thought you were different", "CHALLENGE variant. Step 3. Still no → Step 4 (Downgrade).", PRICE_FILL),
    ("PRICE 4a Downgrade", "okay fine... just for you I'll do it for [lower price] 💗 but only because this convo has been different", "DOWNGRADE. ONE TIME ONLY. Reduce 20-30%. Still no → Step 5.", PRICE_FILL),
    ("PRICE 4b Downgrade", "you know what... [lower price] because you've been making me feel some type of way 💗 don't tell anyone", "DOWNGRADE variant. ONE TIME. Reduce 20-30%. Still no → Step 5.", PRICE_FILL),
    ("PRICE 5a Seed+Return", "it's okay baby 😊 let's just keep talking... I'm still thinking about you anyway", "SEED+RETURN. Accept. Continue GFE. Retry with different content later.", PRICE_FILL),
    ("PRICE 5b Seed+Return", "no worries 💗 we don't have to, I like talking to you regardless", "SEED+RETURN variant. Accept and protect relationship.", PRICE_FILL),

    # ── OBJ-DISCOUNT ──
    ("DISCOUNT 1a Firmness", "haha you trying to negotiate with me? 😏 baby this isn't a market... this is worth every penny trust me", "FIRMNESS. 1st try when he asks discount. Still pushing → Step 2.", DISCOUNT_FILL),
    ("DISCOUNT 1b Firmness", "a discount? 😏 babe do I look like I'm on sale?", "FIRMNESS variant. Still pushing → Step 2 (Challenge).", DISCOUNT_FILL),
    ("DISCOUNT 2a Challenge", "I don't do discounts... I only share this with guys who really appreciate what they're getting 😏", "CHALLENGE WORTHINESS. Step 2. Still pushing → Step 3 (Concession).", DISCOUNT_FILL),
    ("DISCOUNT 2b Challenge", "the guys who value what I do don't ask for discounts 😏 just saying", "CHALLENGE variant. Step 2. Still pushing → Step 3.", DISCOUNT_FILL),
    ("DISCOUNT 3a Concession", "okay you know what... [lower price] just for you 💗 but seriously don't tell anyone, this is our thing", "EXCLUSIVE CONCESSION. ONE TIME. Reduce 20-30%. Frame as secret.", DISCOUNT_FILL),
    ("DISCOUNT 3b Concession", "fine... [lower price] but ONLY because I like you 💗 this is the one and only time", "CONCESSION variant. ONE TIME. Reduce 20-30%.", DISCOUNT_FILL),
    ("DISCOUNT 4a Takeaway", "honestly? if you don't want it that's fine 😏 I'll keep it for myself... or maybe there's someone else who's been asking", "TAKEAWAY. Final step. Jealousy + loss aversion.", DISCOUNT_FILL),
    ("DISCOUNT 4b Takeaway", "okay 😏 I'll save it for someone who really wants it then", "TAKEAWAY variant. Final step.", DISCOUNT_FILL),

    # ── OBJ-FREE ──
    ("FREE 1a Reminder", "baby I already sent you one for free remember? 😏 and this one is even crazier... you have no idea", "REMINDER+TEASE. 1st try. PPV0 was free. Still wants free → Step 2.", FREE_FILL),
    ("FREE 1b Reminder", "you already got a free one 🙈 this one is on another level though, trust me", "REMINDER variant. Still wants free → Step 2 (Challenge).", FREE_FILL),
    ("FREE 2a Challenge", "free? 😏 I don't just give this to anyone... you have to earn the best stuff", "CHALLENGE. Step 2. Buying = earning. Still free → Step 3.", FREE_FILL),
    ("FREE 2b Challenge", "haha free 😏 you think the best things in life are free? not this one baby", "CHALLENGE variant. Still wants free → Step 3 (Guilt).", FREE_FILL),
    ("FREE 3a Guilt", "I literally just spent time recording this because of what YOU said to me 🥺 this was for you, not just random content", "GUILT/INVESTMENT. Step 3. Reciprocity. Still no → Step 4.", FREE_FILL),
    ("FREE 3b Guilt", "I made this because of you... like specifically because of our convo 🥺 it took time and I did it for YOU", "GUILT variant. Reciprocity. Still no → Step 4.", FREE_FILL),
    ("FREE 4a Seed+Return", "it's fine baby 😊 I'm not going anywhere... let's just keep talking, I like you regardless", "SEED+RETURN. Accept. Continue GFE. Retry later.", FREE_FILL),
    ("FREE 4b Seed+Return", "no pressure at all 💗 I'm just happy talking to you honestly", "SEED+RETURN variant. Protect relationship.", FREE_FILL),

    # ── OBJ-NOMONEY ──
    ("NOMONEY 1a Empathy", "aww I totally get it baby 💗 no pressure at all okay?", "EMPATHY. Disarm. Zero judgment. Still engaged → Step 2.", NOMONEY_FILL),
    ("NOMONEY 1b Empathy", "that's okay 💗 don't even worry about it, seriously", "EMPATHY variant. Still engaged → Step 2 (Test).", NOMONEY_FILL),
    ("NOMONEY 2a Test", "not even like [small amount]? 🥺 I really really want you to see this one", "TEST SINCERITY. Suggest $3-5. If says no = likely real. → Step 3.", NOMONEY_FILL),
    ("NOMONEY 2b Test", "what about just [small amount]? 🙈 I really don't want you to miss this", "TEST variant. Suggest $3-5. Still no → Step 3.", NOMONEY_FILL),
    ("NOMONEY 3a PayWhatYouWant", "okay send me whatever you can baby... even a tiny amount 💗 I just need you to see what you made me do", "PAY WHAT YOU WANT. Any amount breaks barrier. Still no → Step 4.", NOMONEY_FILL),
    ("NOMONEY 3b PayWhatYouWant", "just send whatever feels right 🥺 even $1... I just can't keep this from you", "PAY WHAT YOU WANT variant. Still no → Step 4.", NOMONEY_FILL),
    ("NOMONEY 4a Protect", "honestly it's okay 😊 I like talking to you, money or no money... you make me feel things 💗", "PROTECT RELATIONSHIP. Accept. GFE. Tag for re-engagement.", NOMONEY_FILL),
    ("NOMONEY 4b Protect", "baby it's totally fine 💗 you being here is what matters to me", "PROTECT variant. Continue GFE.", NOMONEY_FILL),

    # ── OBJ-NOPPV ──
    ("NOPPV 1a Accept", "that's totally okay 😊 I'm not trying to sell you anything, I just like talking to you", "ACCEPT. Disarm identity stance. Then continue sexting 4-5 msgs.", NOPPV_FILL),
    ("NOPPV 1b Accept", "no worries at all 😊 I don't care about that, I'm just enjoying this convo", "ACCEPT variant. Then continue sexting to build desire.", NOPPV_FILL),
    ("NOPPV 3a EmotionalReframe", "okay but this isn't about money... I just need you to see what you're doing to me rn 🥵 I've never reacted like this to someone", "EMOTIONAL REFRAME. After 4-5 sexting msgs. Still no → Step 4.", NOPPV_FILL),
    ("NOPPV 3b EmotionalReframe", "forget about money for a sec... I just want to share this moment with you 🥵 what you're making me feel is real", "EMOTIONAL REFRAME variant. Still no → Step 4.", NOPPV_FILL),
    ("NOPPV 4a PayWhatYouWant", "just send whatever you want, even $1 🙈 I just can't keep this to myself... you need to see it", "PAY WHAT YOU WANT. $1 breaks identity. He becomes a buyer.", NOPPV_FILL),
    ("NOPPV 4b PayWhatYouWant", "literally send me anything, even the smallest amount 🥺 I need you to see what you did to me", "PAY WHAT YOU WANT variant.", NOPPV_FILL),

    # ── OBJ-CARD ──
    ("CARD 1a Retry", "aww nooo 😩 that happens sometimes baby, try again it usually works the second time", "RETRY. Normalize the issue. Keep energy up.", CARD_FILL),
    ("CARD 1b Retry", "ugh that's so annoying 😩 it happens a lot, try one more time baby", "RETRY variant.", CARD_FILL),
    ("CARD 2a Alternative", "maybe try a different card? 💗 I really don't want you to miss this", "ALTERNATIVE. Practical solution. Still fails → Step 3.", CARD_FILL),
    ("CARD 2b Alternative", "do you have another card you can try? 🥺 I really want you to see this", "ALTERNATIVE variant. Still fails → Step 3.", CARD_FILL),
    ("CARD 3a Urgency", "I really want you to see this before I change my mind 😏 I don't keep stuff like this around forever", "SOFT URGENCY. Don't push hard on tech issue. FOMO + patience.", CARD_FILL),
    ("CARD 3b Urgency", "baby figure it out soon 🙈 I'm in this mood rn and I don't know how long it'll last", "SOFT URGENCY variant.", CARD_FILL),
]

add_rows(ws1, obj_data)

# ═══════════════════════════════════════════════════════
# SHEET 2: ResHandling
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("ResHandling")
setup_sheet(ws2)

res_data = [
    # ── RES-NOSEX ──
    ("NOSEX 1a Respect", "haha okay okay 😊 I got a little carried away, you're just too easy to talk to", "RESPECT. Never force. She 'got carried away'. Still no → Step 2.", RES_FILL),
    ("NOSEX 1b Respect", "oops 😅 sorry I got ahead of myself haha, it's your fault for being so fun to talk to", "RESPECT variant. Still no → Step 2 (Subtle Tension).", RES_FILL),
    ("NOSEX 2a SubtleTension", "so tell me more about you... what do you do when you're not making girls on the internet blush? 😏", "SUBTLE TENSION. Flirt without explicit. Keep attraction alive.", RES_FILL),
    ("NOSEX 2b SubtleTension", "okay okay new topic 😊 but first... what's the most adventurous thing you've ever done? 😏", "SUBTLE TENSION variant. Playful, not sexual.", RES_FILL),
    ("NOSEX 3a ReAttempt", "okay but like... I can't help it 🙈 there's something about you that's messing with my head rn", "NATURAL RE-ATTEMPT. She 'can't help herself'. Still no → Step 4.", RES_FILL),
    ("NOSEX 3b ReAttempt", "I'm trying to behave but you're making it really hard 🙈 there's something about you", "RE-ATTEMPT variant. Still no → Step 4 (Accept).", RES_FILL),
    ("NOSEX 4a Accept", "alright I'll behave... for now 😏 no promises though haha", "ACCEPT. Door stays open. Continue GFE. May convert next session.", RES_FILL),
    ("NOSEX 4b Accept", "okay fine I'll stop 😊 but don't blame me if it happens again later 😏", "ACCEPT variant. Keep door open.", RES_FILL),

    # ── RES-OFFTOPIC ──
    ("OFFTOPIC 1a Acknowledge", "haha wait that's actually funny 😂", "ACKNOWLEDGE. React to what he said. Adapt msg to his topic.", RES_FILL),
    ("OFFTOPIC 1b Acknowledge", "omg 😂 okay that's random but I love it", "ACKNOWLEDGE variant. Validate then redirect.", RES_FILL),
    ("OFFTOPIC 2a Redirect", "but hold on you totally distracted me 😏 I was about to tell you something and now I lost my train of thought...", "REDIRECT WITH HOOK. Creates curiosity. → Step 3 to retake.", RES_FILL),
    ("OFFTOPIC 2b Redirect", "wait no stop 😏 you're distracting me from what I was going to say...", "REDIRECT variant. → Step 3.", RES_FILL),
    ("OFFTOPIC 3a Retake", "okay wait no I remember now 😏 so like I was saying...", "RETAKE CONTROL. Continue with next msg from main journey script.", RES_FILL),
    ("OFFTOPIC 3b Retake", "OKAY focus 😂 where was I... oh yeah 😏", "RETAKE variant. Resume main script flow.", RES_FILL),

    # ── RES-REAL ──
    ("REAL 1a Humor", "lol do I sound like a robot to you? 😂 beep boop... send $5 for human verification 🤖 haha I'm kidding", "HUMOR. Disarms suspicion. A bot wouldn't joke about it.", RES_FILL),
    ("REAL 1b Humor", "wait you think I'm fake?? 😂 that's actually the funniest thing anyone's said to me today", "HUMOR variant. Still doubts → Step 2 (Challenge).", RES_FILL),
    ("REAL 2a Challenge", "ask me anything 😏 literally anything about me, my life, whatever you want. I'm an open book", "CHALLENGE. Invert power. She has nothing to hide.", RES_FILL),
    ("REAL 2b Challenge", "test me 😏 ask me something only a real person would know. go ahead", "CHALLENGE variant. He tests = he's engaged.", RES_FILL),
    ("REAL 3a Grounding", "I get it, there's a lot of fake stuff on here 💗 but you felt something in this convo right? because I did... and that's real", "EMOTIONAL GROUNDING. Anchor on shared experience.", RES_FILL),
    ("REAL 3b Grounding", "I know there's a lot of bots and stuff on here 💗 but what we've been talking about... that felt real to me. didn't it feel real to you?", "GROUNDING variant. Reference emotional connection.", RES_FILL),

    # ── RES-VOICE ──
    ("VOICE 1a Dodge", "haha maybe one day if you're lucky 😏 but not yet... I'm shy about that stuff", "PLAYFUL DODGE. 'Maybe one day' = door not closed. Still asks → Step 2.", RES_FILL),
    ("VOICE 1b Dodge", "hmmm maybe 😏 but you gotta earn that first haha", "DODGE variant. Still asks → Step 2 (Redirect).", RES_FILL),
    ("VOICE 2a Redirect", "I have something wayyy better for you though 🙈 trust me you'll forget you even asked", "REDIRECT to content. Still insists → Step 3.", RES_FILL),
    ("VOICE 2b Redirect", "how about instead of a call, I show you something that'll blow your mind? 😈", "REDIRECT variant. Still insists → Step 3 (Firm).", RES_FILL),
    ("VOICE 3a Firm", "I don't really do that on here baby 💗 but honestly what I'm about to show you is better than any call... you'll see 😈", "FIRM BOUNDARY. Clear but redirects to value.", RES_FILL),
    ("VOICE 3b Firm", "that's not something I do on here 💗 but baby what I have for you is wayyy better than my voice, trust me 😈", "FIRM variant.", RES_FILL),

    # ── RES-CUSTOM-YES ──
    ("CUSTOMYES 1a Tease", "mmm you want that? 😏 I might have something... actually I definitely have something 🙈", "TEASE. Build anticipation. Don't give instantly.", RES_FILL),
    ("CUSTOMYES 1b Tease", "ohhh 😏 you have good taste... I think I have exactly what you're looking for", "TEASE variant. → Step 2 (Price).", RES_FILL),
    ("CUSTOMYES 2a Price", "okay I have exactly what you're thinking of 😈 you're gonna lose your mind... [price]", "PRICE. Set price based on content. Premium for specific requests.", RES_FILL),
    ("CUSTOMYES 2b Price", "I actually made something just like that 🙈 [price] and it's worth every penny baby", "PRICE variant. Set appropriate price.", RES_FILL),
    ("CUSTOMYES 3a Close", "trust me baby you won't regret it 😈 I made this one special", "CLOSE. Reinforce decision. Send content.", RES_FILL),
    ("CUSTOMYES 3b Close", "you're not gonna be able to stop watching this one 🥵", "CLOSE variant.", RES_FILL),

    # ── RES-CUSTOM-NO ──
    ("CUSTOMNO 1a Redirect", "I don't have exactly that but honestly I have something that'll make you forget you even asked 😏", "REDIRECT. Never say 'no' flat out. Offer better.", RES_FILL),
    ("CUSTOMNO 1b Redirect", "hmm I don't have that specific thing but I have something you're gonna love even more 😈", "REDIRECT variant. → Step 2.", RES_FILL),
    ("CUSTOMNO 2a Alternative", "actually what I have might be even crazier 🙈 and literally no one else has seen it yet", "ALTERNATIVE + FOMO. Exclusivity. → Step 3.", RES_FILL),
    ("CUSTOMNO 2b Alternative", "what I DO have is something no one has ever seen 🙈 and I think it's even better than what you asked for", "ALTERNATIVE variant. → Step 3.", RES_FILL),
    ("CUSTOMNO 3a Close", "trust me baby... I know what you need better than you do 😈", "CONFIDENCE CLOSE. Model leads.", RES_FILL),
    ("CUSTOMNO 3b Close", "just trust me on this one 😈 you'll thank me later", "CLOSE variant.", RES_FILL),

    # ── RES-DONE ──
    ("DONE 1a Validate", "fuck that's so hot 🥵 you came because of me??", "VALIDATE. Celebrate it. Feed ego. → Step 2 (Rescue).", RES_FILL),
    ("DONE 1b Validate", "omg 🥵🥵 already?? that's so hot baby", "VALIDATE variant. → Step 2.", RES_FILL),
    ("DONE 2a Rescue", "but baby I haven't finished yet 🥺 don't you wanna watch me cum too?", "RESCUE. Flip to HER. Reciprocity + ego. Still no → Step 3.", RES_FILL),
    ("DONE 2b Rescue", "wait but I'm not done yet 🥺 you're gonna leave me like this?", "RESCUE variant. Still no → Step 3 (Seed).", RES_FILL),
    ("DONE 3a Seed", "okay but next time you have to wait for me 😏 I have something insane planned for round 2", "SEED NEXT SESSION. Plants anticipation for return.", RES_FILL),
    ("DONE 3b Seed", "next time you HAVE to hold it 😏 because what I have planned for us next time is way crazier", "SEED variant.", RES_FILL),
]

add_rows(ws2, res_data)

# ═══════════════════════════════════════════════════════
# SHEET 3: Situational
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Situational")
setup_sheet(ws3)

sit_data = [
    # ── SIT-CUM ──
    ("CUM A1 Edge", "don't cum yet baby... I'm not done with you 😈", "EDGE. Use when more PPVs left. Slows him down.", SIT_FILL),
    ("CUM A2 Edge", "hold it 🥵 not yet... I need you to last a little longer for me", "EDGE variant. More PPVs to sell.", SIT_FILL),
    ("CUM B1 Sync", "I'm so close too 🥵 cum with me baby... but you need to see this first 💦", "SYNC. Use at FINAL PPV. Time orgasm with last purchase.", SIT_FILL),
    ("CUM B2 Sync", "wait for me 🥵 I want us to finish together... open this first 💦", "SYNC variant. Final PPV timing.", SIT_FILL),
    ("CUM C1 TeaseDelay", "hold it... I want you to wait until you see what I'm about to send 😏 trust me it's worth it", "TEASE DELAY. Next PPV is about to be sent.", SIT_FILL),
    ("CUM C2 TeaseDelay", "don't you dare finish before you see this 😏 trust me you want to wait", "TEASE DELAY variant.", SIT_FILL),

    # ── SIT-DICKPIC ──
    ("DICKPIC A1 Sexting", "fuck 🥵🥵 okay that's... wow. you have no idea what that just did to me", "DURING SEXTING. React positively. Then leverage for PPV.", SIT_FILL),
    ("DICKPIC A2 Sexting", "holy shit 🥵 that is... fuckkk. I need to show you something rn", "DURING SEXTING variant. Leads into next PPV.", SIT_FILL),
    ("DICKPIC B1 Rapport", "omg 🙈 you don't waste time huh 😏 that's actually really hot though ngl", "DURING RAPPORT. Surprised but into it. Fast-track to teasing.", SIT_FILL),
    ("DICKPIC B2 Rapport", "woah 🙈 I wasn't expecting that but... damn 😏", "DURING RAPPORT variant.", SIT_FILL),
    ("DICKPIC C1 Leverage", "you can't just send me that and expect me not to do something about it 🥵 hold on...", "LEVERAGE FOR PPV. His pic 'caused' her to record. WAIT 1-2 min then send PPV.", SIT_FILL),
    ("DICKPIC C2 Leverage", "okay you just made me do something... give me a sec 🥵", "LEVERAGE variant. WAIT 1-2 min then send PPV.", SIT_FILL),

    # ── SIT-HORNY (pool) ──
    ("HORNY 1", "fuckkk 🥵", "MID-SEXTING BOOSTER. Pick any from this pool to maintain energy.", SIT_FILL),
    ("HORNY 2", "I'm so wet rn because of you", "BOOSTER. Use between script messages for extra intensity.", SIT_FILL),
    ("HORNY 3", "don't stop", "BOOSTER. Micro message. Keeps momentum.", SIT_FILL),
    ("HORNY 4", "you have no idea what you're doing to me", "BOOSTER. Ego feed + sexual tension.", SIT_FILL),
    ("HORNY 5", "I literally can't think straight rn", "BOOSTER. Shows she's overwhelmed by him.", SIT_FILL),
    ("HORNY 6", "my hands are shaking 🥵", "BOOSTER. Physical vulnerability detail.", SIT_FILL),
    ("HORNY 7", "more...", "BOOSTER. Ultra micro. Desperate energy.", SIT_FILL),
    ("HORNY 8", "I should be packing for my trip but I can't move rn 🥵", "BOOSTER. Putri personality touch (travel/adventure).", SIT_FILL),
]

add_rows(ws3, sit_data)

# ── SAVE ──
output_path = r"c:\Users\34683\CW-ScriptManager\putri\Putri_OBJ_RES_SIT_Infloww.xlsx"
wb.save(output_path)

sheets = [ws.title for ws in wb.worksheets]
counts = [ws.max_row - 1 for ws in wb.worksheets]
print(f"Saved: {output_path}")
print(f"Sheets: {sheets}")
print(f"Scripts per sheet: {dict(zip(sheets, counts))}")
print(f"Total scripts: {sum(counts)}")
