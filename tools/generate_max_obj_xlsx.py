"""
Generate Max OBJ/RES/SIT XLSX for Infloww import.
Voice: Alpha, dominant, masculine. No "baby/babe/honey". Uses direct language.
Audience: Gay male (dating app traffic).
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(bold=True, color="e6edf3", size=11)
OBJ_FILL = PatternFill("solid", fgColor="2d1a1a")
RES_FILL = PatternFill("solid", fgColor="1f1a2d")
SIT_FILL = PatternFill("solid", fgColor="1a2a1a")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="30363d"))

def make_sheet(title, rows, fill):
    ws = wb.create_sheet(title)
    ws.append(["Name", "Text", "Note", "*Guidelines"])
    for c in ["A","B","C","D"]:
        ws[f"{c}1"].fill = HEADER_FILL
        ws[f"{c}1"].font = HEADER_FONT
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 25
    for name, text, note in reversed(rows):
        ws.append([name, text, note, ""])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP
            cell.border = THIN
            if fill:
                cell.fill = fill

# ═══════════════════════════════════════════
#  A. OBJECTION HANDLING
# ═══════════════════════════════════════════

make_sheet("price1", [
    ("Step1 Reframe", "bro that's less than a protein shake and I promise this hits harder", "REFRAME. Compare to cheap item. Wait for response. Still no -> Step 2."),
    ("Step2 FOMO", "I'm only feeling like this because of you rn, no idea when that's gonna happen again", "FOMO. Temporal urgency. Still no -> Step 3."),
    ("Step3 Challenge", "maybe you're not ready for what I just did in this one", "CHALLENGE. Ego. Still no -> Step 4."),
    ("Step4 Downgrade", "alright look, [lower price] just for you because this convo has been different", "DOWNGRADE. Reduce 20-30%. ONE TIME. Still no -> Step 5."),
    ("Step5 Seed", "it's cool man, let's just keep talking... I'm still thinking about you anyway", "SEED. Accept. Continue GFE. Retry later."),
], OBJ_FILL)

make_sheet("price2", [
    ("Step1 Reframe", "that's like what you'd spend on a coffee, except this is gonna keep you up way longer", "REFRAME. Still no -> Step 2."),
    ("Step2 FOMO", "this mood? it's not gonna last forever and I want you to be the one who sees it", "FOMO. Still no -> Step 3."),
    ("Step3 Challenge", "honestly most guys can't handle what I just recorded, thought you were different", "CHALLENGE. Still no -> Step 4."),
    ("Step4 Downgrade", "you know what... [lower price] because you've been making me feel some type of way, keep that between us", "DOWNGRADE. ONE TIME. Still no -> Step 5."),
    ("Step5 Seed", "no stress, I like talking to you regardless", "SEED. Protect relationship."),
], OBJ_FILL)

make_sheet("discount1", [
    ("Step1 Firmness", "haha you trying to negotiate with me? this isn't a negotiation man, it's worth every cent", "FIRMNESS. Playful. Still pushing -> Step 2."),
    ("Step2 Challenge", "I don't do discounts... I only share this with guys who actually appreciate what they're getting", "CHALLENGE. Worthiness. Still pushing -> Step 3."),
    ("Step3 Concession", "alright... [lower price] just for you but don't tell anyone, this stays between us", "CONCESSION. ONE TIME. Still no -> Step 4."),
    ("Step4 Takeaway", "if you don't want it that's cool, I'll keep it for myself... or maybe someone else has been asking", "TAKEAWAY. Jealousy. Final step."),
], OBJ_FILL)

make_sheet("discount2", [
    ("Step1 Firmness", "a discount? do I look like I'm on sale?", "FIRMNESS. Short + confident. Still pushing -> Step 2."),
    ("Step2 Challenge", "the guys who appreciate what I do don't ask for discounts, just saying", "CHALLENGE. Still pushing -> Step 3."),
    ("Step3 Concession", "fine... [lower price] but ONLY because I like you, one time thing", "CONCESSION. ONE TIME. Still no -> Step 4."),
    ("Step4 Takeaway", "alright I'll save it for someone who actually wants it then", "TAKEAWAY. Final."),
], OBJ_FILL)

make_sheet("free1", [
    ("Step1 Reminder", "I already sent you one for free remember? and this one is way crazier... you have no idea", "REMINDER. PPV0 was free. Still wants free -> Step 2."),
    ("Step2 Challenge", "free? nah I don't just give this to anyone... you gotta earn the good stuff", "CHALLENGE. Buying = earning. Still free -> Step 3."),
    ("Step3 Guilt", "I literally just recorded this because of what YOU said to me, this wasn't just random content man", "GUILT. Reciprocity. Still no -> Step 4."),
    ("Step4 Seed", "it's cool, I'm not going anywhere... let's just keep talking, I like this", "SEED. Protect relationship."),
], OBJ_FILL)

make_sheet("free2", [
    ("Step1 Reminder", "you already got a free one, this one is on another level trust me", "REMINDER. Still wants free -> Step 2."),
    ("Step2 Challenge", "free? you think the best things in life are free? not this one", "CHALLENGE. Still free -> Step 3."),
    ("Step3 Guilt", "I did this because of you... like specifically because of our convo, that took effort and I did it for YOU", "GUILT. Still no -> Step 4."),
    ("Step4 Seed", "no pressure at all, I'm just enjoying talking to you honestly", "SEED."),
], OBJ_FILL)

make_sheet("nomoney1", [
    ("Step1 Empathy", "hey I totally get it, no pressure at all okay?", "EMPATHY. Zero judgment. Still engaged -> Step 2."),
    ("Step2 Test", "not even like [small amount]? I really want you to see this one", "TEST. Suggest $3-5. Still no -> Step 3."),
    ("Step3 PWYW", "just send whatever you can man... even a tiny amount, I just need you to see what you made me do", "PAY WHAT YOU WANT. Still no -> Step 4."),
    ("Step4 Protect", "honestly it's fine, I like talking to you money or not... you do something to me", "PROTECT. GFE. Tag for re-engagement."),
], OBJ_FILL)

make_sheet("nomoney2", [
    ("Step1 Empathy", "that's fine, seriously don't worry about it", "EMPATHY. Still engaged -> Step 2."),
    ("Step2 Test", "what about just [small amount]? I really don't want you to miss this", "TEST. Still no -> Step 3."),
    ("Step3 PWYW", "send whatever feels right, even $1... I just can't keep this from you", "PWYW. Still no -> Step 4."),
    ("Step4 Protect", "it's totally cool, you being here is what matters", "PROTECT. Continue conversation."),
], OBJ_FILL)

make_sheet("noppv1", [
    ("Step1 Accept", "that's totally fine I'm not trying to sell you anything, I just like talking to you", "ACCEPT. Disarm. THEN: continue sexting 4-5 msgs before Step 2."),
    ("Step2 Reframe", "look this isn't about money... I just need you to see what you're doing to me rn, I don't react like this to people", "EMOTIONAL REFRAME. After sexting buildup. Still no -> Step 3."),
    ("Step3 PWYW", "just send whatever you want, even $1, I can't keep this to myself... you need to see it", "PAY WHAT YOU WANT. $1 breaks identity."),
], OBJ_FILL)

make_sheet("noppv2", [
    ("Step1 Accept", "no worries at all, I don't care about that I'm just enjoying this", "ACCEPT. Disarm. THEN: continue sexting 4-5 msgs before Step 2."),
    ("Step2 Reframe", "forget about money for a sec... I just want to share this with you, what you're making me feel is real", "REFRAME. Still no -> Step 3."),
    ("Step3 PWYW", "send me anything, even the smallest amount, I need you to see what you did to me", "PWYW."),
], OBJ_FILL)

make_sheet("card1", [
    ("Step1 Retry", "ahh that sucks, happens sometimes though try again it usually works the second time", "RETRY. Normalize. Still fails -> Step 2."),
    ("Step2 AltCard", "try a different card? I really don't want you to miss this", "ALTERNATIVE. Still fails -> Step 3."),
    ("Step3 Urgency", "figure it out soon man, I'm in this mood rn and I don't know how long it's gonna last", "SOFT URGENCY. FOMO."),
], OBJ_FILL)

make_sheet("card2", [
    ("Step1 Retry", "damn that's annoying, it happens a lot just try one more time", "RETRY. Still fails -> Step 2."),
    ("Step2 AltCard", "do you have another card you can try? I really want you to see this", "ALTERNATIVE. Still fails -> Step 3."),
    ("Step3 Urgency", "I want you to see this before I change my mind, I don't keep stuff like this around forever", "URGENCY."),
], OBJ_FILL)

# ═══════════════════════════════════════════
#  B. RESISTANCE HANDLING
# ═══════════════════════════════════════════

make_sheet("nosex1", [
    ("Step1 Respect", "haha alright alright I got a little carried away, you're just too easy to talk to", "RESPECT. Never force. Still no -> Step 2."),
    ("Step2 Subtle", "so tell me more about you... what do you do when you're not making dudes on the internet lose their minds?", "SUBTLE TENSION. Flirt without explicit. -> Step 3 later."),
    ("Step3 ReAttempt", "look I can't help it, there's something about you that's messing with my head rn", "RE-ATTEMPT. Natural. Still no -> Step 4."),
    ("Step4 Accept", "alright I'll chill... for now, no promises though haha", "ACCEPT. Door stays open."),
], RES_FILL)

make_sheet("nosex2", [
    ("Step1 Respect", "my bad I got ahead of myself haha, it's your fault for being so fun to talk to", "RESPECT. Still no -> Step 2."),
    ("Step2 Subtle", "okay new topic, but first... what's the most wild thing you've ever done?", "SUBTLE TENSION. -> Step 3 later."),
    ("Step3 ReAttempt", "I'm trying to behave but you're making it really hard man, there's something about you", "RE-ATTEMPT. Still no -> Step 4."),
    ("Step4 Accept", "fine I'll stop but don't blame me if it happens again later", "ACCEPT. Door open."),
], RES_FILL)

make_sheet("offtopic1", [
    ("Step1 Acknowledge", "haha wait that's actually funny", "ACKNOWLEDGE. Adapt to what he actually said. -> Step 2."),
    ("Step2 Redirect", "but hold on you totally distracted me, I was about to tell you something and now I lost my train of thought...", "REDIRECT. Curiosity hook. -> Step 3."),
    ("Step3 Retake", "okay wait no I remember now, so like I was saying...", "RETAKE. Continue with next msg from main journey."),
], RES_FILL)

make_sheet("offtopic2", [
    ("Step1 Acknowledge", "lol okay that's random but I like it", "ACKNOWLEDGE. Adapt. -> Step 2."),
    ("Step2 Redirect", "wait no stop you're distracting me from what I was gonna say...", "REDIRECT. -> Step 3."),
    ("Step3 Retake", "OKAY focus, where was I... oh yeah", "RETAKE. Resume main script."),
], RES_FILL)

make_sheet("real1", [
    ("Step1 Humor", "lol do I sound like a bot to you? beep boop... send $5 for human verification haha I'm kidding", "HUMOR. Disarm. Still doubts -> Step 2."),
    ("Step2 Challenge", "ask me anything, literally anything about me my life whatever. I'm an open book", "CHALLENGE. He tests = engaged. Still doubts -> Step 3."),
    ("Step3 Grounding", "I get it there's a lot of fake stuff on here, but you felt something in this convo right? because I did... and that's real", "GROUNDING. Anchor on emotion."),
], RES_FILL)

make_sheet("real2", [
    ("Step1 Humor", "wait you think I'm fake?? that's actually the funniest thing anyone's said to me today", "HUMOR. Still doubts -> Step 2."),
    ("Step2 Challenge", "test me, ask me something only a real person would know. go ahead", "CHALLENGE. Still doubts -> Step 3."),
    ("Step3 Grounding", "I know there's a lot of bots on here but what we've been talking about... that felt real to me. didn't it feel real to you?", "GROUNDING."),
], RES_FILL)

make_sheet("voice1", [
    ("Step1 Dodge", "haha maybe one day if you earn it but not yet... I'm private about that stuff", "DODGE. Door not closed. Still asks -> Step 2."),
    ("Step2 Redirect", "I have something wayyy better for you though, trust me you'll forget you even asked", "REDIRECT. Still insists -> Step 3."),
    ("Step3 Firm", "I don't do that on here but what I'm about to show you is better than any call... you'll see", "FIRM BOUNDARY. Redirect to value."),
], RES_FILL)

make_sheet("voice2", [
    ("Step1 Dodge", "hmmm maybe but you gotta earn that first haha", "DODGE. Still asks -> Step 2."),
    ("Step2 Redirect", "how about instead of a call I show you something that'll blow your mind?", "REDIRECT. Still insists -> Step 3."),
    ("Step3 Firm", "that's not something I do on here but what I have for you is wayyy better than my voice, trust me", "FIRM."),
], RES_FILL)

make_sheet("customyes1", [
    ("Step1 Tease", "you want that? I might have something... actually I definitely have something", "TEASE. Build anticipation. -> Step 2."),
    ("Step2 Price", "I have exactly what you're thinking of, you're gonna lose it... [price]", "PRICE. Set based on content."),
    ("Step3 Close", "trust me you won't regret it, I made this one special", "CLOSE."),
], RES_FILL)

make_sheet("customyes2", [
    ("Step1 Tease", "oh you have good taste... I think I have exactly what you're looking for", "TEASE. -> Step 2."),
    ("Step2 Price", "I actually did something just like that, [price] and it's worth every cent", "PRICE."),
    ("Step3 Close", "you're not gonna be able to stop watching this one", "CLOSE."),
], RES_FILL)

make_sheet("customno1", [
    ("Step1 Redirect", "I don't have exactly that but honestly I have something that'll make you forget you even asked", "REDIRECT. -> Step 2."),
    ("Step2 Alternative", "actually what I have might be even crazier and literally no one else has seen it yet", "ALTERNATIVE + FOMO. -> Step 3."),
    ("Step3 Close", "trust me... I know what you need better than you do", "CONFIDENCE CLOSE."),
], RES_FILL)

make_sheet("customno2", [
    ("Step1 Redirect", "I don't have that specific thing but I have something you're gonna like even more", "REDIRECT. -> Step 2."),
    ("Step2 Alternative", "what I DO have is something no one has ever seen and I think it's even better than what you asked for", "ALTERNATIVE. -> Step 3."),
    ("Step3 Close", "just trust me on this one, you'll thank me later", "CLOSE."),
], RES_FILL)

make_sheet("done1", [
    ("Step1 Validate", "fuck that's hot, you came because of me??", "VALIDATE. Feed ego. -> Step 2."),
    ("Step2 Rescue", "but I haven't finished yet, don't you wanna watch me cum too?", "RESCUE. Flip to HIM. Still no -> Step 3."),
    ("Step3 Seed", "okay but next time you have to wait for me, I have something insane planned for round 2", "SEED. Anticipation."),
], RES_FILL)

make_sheet("done2", [
    ("Step1 Validate", "already?? damn that's hot", "VALIDATE. -> Step 2."),
    ("Step2 Rescue", "wait but I'm not done yet, you're gonna leave me like this?", "RESCUE. Still no -> Step 3."),
    ("Step3 Seed", "next time you HAVE to hold it because what I have planned for us next time is way crazier", "SEED."),
], RES_FILL)

# ═══════════════════════════════════════════
#  C. SITUATIONAL — standalone
# ═══════════════════════════════════════════

make_sheet("cumcontrol", [
    ("edge1", "don't cum yet... I'm not done with you", "EDGE. More PPVs left."),
    ("edge2", "hold it, not yet... I need you to last a little longer for me", "EDGE variant."),
    ("sync1", "I'm so close too, cum with me... but you need to see this first", "SYNC. Final PPV."),
    ("sync2", "wait for me, I want us to finish together... open this first", "SYNC variant."),
    ("delay1", "hold it... I want you to wait until you see what I'm about to send, trust me it's worth it", "DELAY. Next PPV coming."),
    ("delay2", "don't you dare finish before you see this, trust me you want to wait", "DELAY variant."),
], SIT_FILL)

make_sheet("dickpic", [
    ("dpsext1", "fuck okay that's... damn. you have no idea what that just did to me", "DURING SEXTING. React + leverage for PPV."),
    ("dpsext2", "holy shit that is... fuck. I need to show you something rn", "DURING SEXTING variant."),
    ("dprapport1", "damn you don't waste time huh, that's actually really hot though ngl", "DURING RAPPORT. Fast-track to teasing."),
    ("dprapport2", "woah I wasn't expecting that but... damn", "DURING RAPPORT variant."),
    ("dpppv1", "you can't just send me that and expect me not to do something about it, hold on...", "LEVERAGE. WAIT 1-2 min then send PPV."),
    ("dpppv2", "okay you just made me do something... give me a sec", "LEVERAGE variant. WAIT 1-2 min."),
], SIT_FILL)

make_sheet("boosters", [
    ("h1", "fuckkk", "MID-SEXTING BOOSTER."),
    ("h2", "I'm so hard rn because of you", "BOOSTER. Ego feed."),
    ("h3", "don't stop", "BOOSTER. Micro."),
    ("h4", "you have no idea what you're doing to me", "BOOSTER."),
    ("h5", "I literally can't think straight rn", "BOOSTER."),
    ("h6", "my hands are shaking", "BOOSTER. Physical."),
    ("h7", "more...", "BOOSTER. Ultra micro."),
    ("h8", "I should be at the gym but I can't move rn", "BOOSTER. Max personality."),
], SIT_FILL)

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

output = r"c:\Users\34683\CW-ScriptManager\max\Max_OBJ_RES_SIT_Infloww.xlsx"
wb.save(output)

sheets = [ws.title for ws in wb.worksheets]
total = sum(ws.max_row - 1 for ws in wb.worksheets)
print(f"Saved: {output}")
print(f"Sheets ({len(sheets)}): {sheets}")
print(f"Total scripts: {total}")
