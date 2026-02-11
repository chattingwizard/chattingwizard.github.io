"""
Merge Journey + OBJ/RES XLSX files into ONE file per model.
Copies all sheets from both files into a single workbook.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
from copy import copy

def copy_sheet(src_ws, dst_wb, title):
    """Copy a sheet from source to destination workbook, preserving styles."""
    dst_ws = dst_wb.create_sheet(title)
    
    # Copy column dimensions
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
    
    # Copy rows with values and styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.number_format = cell.number_format

def merge_files(journey_path, obj_path, output_path):
    """Merge two xlsx files into one."""
    wb_out = openpyxl.Workbook()
    
    # Load journey
    wb_j = openpyxl.load_workbook(journey_path)
    first = True
    for ws in wb_j.worksheets:
        if first:
            # Use the default sheet for the first one
            dst = wb_out.active
            dst.title = ws.title
            for col_letter, col_dim in ws.column_dimensions.items():
                dst.column_dimensions[col_letter].width = col_dim.width
            for row in ws.iter_rows():
                for cell in row:
                    new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = cell.number_format
            first = False
        else:
            copy_sheet(ws, wb_out, ws.title)
    wb_j.close()
    
    # Load OBJ/RES
    wb_o = openpyxl.load_workbook(obj_path)
    for ws in wb_o.worksheets:
        copy_sheet(ws, wb_out, ws.title)
    wb_o.close()
    
    wb_out.save(output_path)
    sheets = [ws.title for ws in wb_out.worksheets]
    total = sum(ws.max_row - 1 for ws in wb_out.worksheets)
    print(f"Saved: {output_path}")
    print(f"Sheets ({len(sheets)}): {sheets}")
    print(f"Total scripts: {total}")
    return output_path

# ── MAX ──
print("=" * 60)
print("MAX")
print("=" * 60)
merge_files(
    r"c:\Users\34683\CW-ScriptManager\max\Max_DatingApp_Journey_Infloww.xlsx",
    r"c:\Users\34683\CW-ScriptManager\max\Max_OBJ_RES_SIT_Infloww.xlsx",
    r"c:\Users\34683\CW-ScriptManager\max\Max_Complete_Infloww.xlsx",
)

# ── PUTRI ──
print()
print("=" * 60)
print("PUTRI")
print("=" * 60)
merge_files(
    r"c:\Users\34683\CW-ScriptManager\putri\Putri_New_Sub_Journey_Infloww.xlsx",
    r"c:\Users\34683\CW-ScriptManager\putri\Putri_OBJ_RES_SIT_Infloww.xlsx",
    r"c:\Users\34683\CW-ScriptManager\putri\Putri_Complete_Infloww.xlsx",
)
