"""
MODEL FACTORY — Master generator for Chatting Wizard
Generates complete package for any model:
  1. Journey XLSX (all sheets)
  2. OBJ/RES/SIT XLSX (29 sheets)
  3. Merged single XLSX
  4. HTML chatter guide
  5. HTML objections guide
  6. Downloads profile photo
  7. Updates dashboard

Usage:
  from model_factory import ModelFactory
  m = ModelFactory(model_config)
  m.generate_all()
"""
import os, sys, json, requests, html as html_mod
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

# ══════════════════════════════════════════
# STYLE CONSTANTS
# ══════════════════════════════════════════
HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PPV_FILL = PatternFill("solid", fgColor="FFF6B26B")
WAIT_FILL = PatternFill("solid", fgColor="FFFFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="FFD9EAD3")
PURPLE_FILL = PatternFill("solid", fgColor="FFD9D2E9")
AC_FILL = PatternFill("solid", fgColor="FFD0E0E3")
BLUE1 = PatternFill("solid", fgColor="FFCFE2F3")
BLUE2 = PatternFill("solid", fgColor="FFE0F7FA")
OBJ_HDR = PatternFill("solid", fgColor="1a1a2e")
OBJ_HDR_FONT = Font(bold=True, color="e6edf3", size=11)
OBJ_FILL = PatternFill("solid", fgColor="2d1a1a")
RES_FILL = PatternFill("solid", fgColor="1f1a2d")
SIT_FILL = PatternFill("solid", fgColor="1a2a1a")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="30363d"))

# ── Protocol descriptions: explains WHEN to use each OBJ/RES/SIT protocol ──
PROTOCOL_DESCRIPTIONS = {
    "price": "Fan says it's too expensive. Reframe value → FOMO → Challenge → Downgrade (once) → Seed.",
    "discount": "Fan asks for a discount. Firmness → Challenge → One-time concession → Takeaway.",
    "free": "Fan wants free content. Remind PPV0 was free → Challenge → Guilt → Seed.",
    "nomoney": "Fan says no money. Empathy → Test sincerity → PWYW → Protect relationship.",
    "noppv": "Fan says 'I don't buy PPVs'. Accept → Keep sexting → Emotional reframe → PWYW.",
    "card": "Card doesn't work. Retry → Alt card → Urgency.",
    "nosex": "Fan doesn't want to go sexual. Respect → Subtle tension → Re-attempt → Accept.",
    "offtopic": "Fan goes off-topic. Acknowledge → Redirect → Retake control.",
    "real": "Fan asks 'are you real?' / 'is this a bot?'. Humor → Challenge → Emotional grounding.",
    "voice": "Fan asks for voice note or video call. Playful dodge → Redirect to content → Firm.",
    "customyes": "Fan asks for content we HAVE. Tease → Price → Close.",
    "customno": "Fan asks for content we DON'T have. Redirect → Alternative + FOMO → Close.",
    "done": "Fan already finished before final PPV. Validate → Rescue → Seed next session.",
    "cumcontrol": "Situational: control orgasm pacing to maximize PPVs sold.",
    "dickpic": "Situational: fan sends dick pic. React positively → Leverage for PPV.",
    "boosters": "Pool of short messages to keep mid-sexting intensity between PPVs.",
}

GUIDELINES = """Character limits
Name: Up to 64 characters
Tag: Up to 32 characters
Text: Up to 1000 characters
Note: Up to 246 characters

Each sheet serves as a tag. Simply replace the sheet name with your desired tag name, and all scripts added to the sheet will be automatically assigned that tag once imported."""


def setup_journey_sheet(ws):
    ws.append(["Name", "Text", "Note", "*Guidelines"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 20


def add_rows_reversed(ws, rows):
    reversed_rows = list(reversed(rows))
    fill_map = {
        "ppv": PPV_FILL, "wait": WAIT_FILL, "rapport": GREEN_FILL,
        "teasing": PURPLE_FILL, "aftercare": AC_FILL
    }
    for idx, (name, text, note, row_type) in enumerate(reversed_rows):
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=text)
        ws.cell(row=row_num, column=3, value=note or "")
        if idx == 0:
            ws.cell(row=row_num, column=4, value=GUIDELINES)
        fill = fill_map.get(row_type)
        if row_type == "sext":
            fill = BLUE1 if idx % 2 == 0 else BLUE2
        if fill:
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).fill = fill
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).alignment = WRAP


def make_obj_sheet(wb, title, rows, fill):
    ws = wb.create_sheet(title)
    ws.append(["Name", "Text", "Note", "*Guidelines"])
    for c in ["A","B","C","D"]:
        ws[f"{c}1"].fill = OBJ_HDR
        ws[f"{c}1"].font = OBJ_HDR_FONT
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 25
    for name, text, note in reversed(rows):
        # Prefix sheet name to ensure unique command names across all sheets
        unique_name = f"{title} {name}" if not name.startswith(title) else name
        # Infloww max 64 chars for Name column
        if len(unique_name) > 64:
            unique_name = unique_name[:64]
        ws.append([unique_name, text, note, ""])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.alignment = WRAP
            cell.border = THIN
            if fill:
                cell.fill = fill


def add_standalone_rows(ws, rows):
    """Add rows without reversing (for Personal Info, Positive Spin)."""
    for idx, (name, text, note) in enumerate(rows):
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=text)
        ws.cell(row=row_num, column=3, value=note or "")
        if idx == 0:
            ws.cell(row=row_num, column=4, value=GUIDELINES)
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).alignment = WRAP


def merge_workbooks(wb_journey, wb_obj, output_path):
    """Merge two workbooks into one."""
    wb_out = openpyxl.Workbook()
    
    first = True
    for wb in [wb_journey, wb_obj]:
        for ws in wb.worksheets:
            if first:
                dst = wb_out.active
                dst.title = ws.title
                first = False
            else:
                dst = wb_out.create_sheet(ws.title)
            
            for col_letter, col_dim in ws.column_dimensions.items():
                dst.column_dimensions[col_letter].width = col_dim.width
            
            for row in ws.iter_rows():
                for cell in row:
                    new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = cell.number_format
    
    wb_out.save(output_path)
    sheets = [ws.title for ws in wb_out.worksheets]
    total = sum(ws.max_row - 1 for ws in wb_out.worksheets)
    return len(sheets), total


def download_photo(model_name, dest_folder):
    """Download profile photo from Airtable."""
    token = os.environ.get("AIRTABLE_TOKEN", "")
    if not token:
        print(f"  [SKIP] No AIRTABLE_TOKEN for photo download")
        return None
    
    base_id = "appy0qGaMEfyDz9LZ"
    table_id = "tbl97sE9V8wbcgjAJ"
    
    formula = f"{{Model Name}} = '{model_name}'"
    r = requests.get(
        f"https://api.airtable.com/v0/{base_id}/{table_id}",
        headers={"Authorization": f"Bearer {token}"},
        params={"filterByFormula": formula, "fields[]": ["Profile Picture"]}
    )
    data = r.json()
    records = data.get("records", [])
    if not records:
        print(f"  [SKIP] No Airtable record for '{model_name}'")
        return None
    
    pics = records[0].get("fields", {}).get("Profile Picture", [])
    if not pics:
        print(f"  [SKIP] No profile picture for '{model_name}'")
        return None
    
    url = pics[0].get("url", "")
    if not url:
        return None
    
    ext = ".jpg"
    if "png" in url.lower():
        ext = ".png"
    elif "jpeg" in url.lower() or "jpg" in url.lower():
        ext = ".jpeg"
    
    dest = os.path.join(dest_folder, f"profile{ext}")
    img_data = requests.get(url).content
    with open(dest, "wb") as f:
        f.write(img_data)
    print(f"  [OK] Photo saved: {dest}")
    return f"profile{ext}"


# ══════════════════════════════════════════
# MAIN FACTORY CLASS
# ══════════════════════════════════════════

class ModelFactory:
    """
    config = {
        "name": "Marco",
        "airtable_name": "Marco",       # Name in Airtable
        "folder": "marco",              # Folder name for web
        "gender": "male",               # male/female
        "traffic": "dating_app",        # dating_app / social_media
        "age": 25,
        "nationality": "Turkish",
        "location": "Texas, USA",
        "page_type": "Paid Page",
        "personality": "Masculine, confident, in control...",
        "voice": "Lowercase. Direct. Short. Confident...",
        "voice_pet_names": "bro, man, dude",
        "voice_never": "baby, babe, honey",
        "interests": ["gym", "meat", "travel", "role-play"],
        "physical": "175cm, 82kg, brown hair, brown eyes, no tattoos",
        "job": "Gym instructor",
        "countries": "USA, Turkey, Spain, Poland, France",
        "key_phrases": ["...", "..."],
        "explicit_level": "full",       # full / non_explicit / soft
        "special_notes": "...",
        
        # Content for scripts (provided per model)
        "journey": [...],               # List of journey messages
        "obj_scripts": {...},           # OBJ/RES/SIT data
        "personal_info": [...],
        "positive_spin": [...],
        "nr_waves": [...],
        # Dating app specific:
        "meetup_redirect": [...],
        "location_handling": [...],
        "re_engagement": [...],
    }
    """
    
    def __init__(self, config):
        self.c = config
        self.base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.web_dir = os.path.join(self.base_dir, self.c["folder"])
        self.scripts_dir = os.path.join(self.base_dir, "scripts", "models", self.c["folder"])
        self._apply_obj_defaults()
        
    def _resolve_gender_key(self):
        """Return normalized gender key: 'female', 'male', or 'male_gay'."""
        gender = self.c.get("gender", "female")
        traffic = self.c.get("traffic", "social_media")
        # Jack Hollywood and similar DL/gay-audience males
        special = self.c.get("special_notes", "") or ""
        if gender == "male" and ("gay" in traffic.lower() or "gay" in special.lower()
                                  or "DL" in special or "dl" in special.lower()):
            return "male_gay"
        elif gender == "male":
            return "male"
        return "female"

    def _apply_obj_defaults(self):
        """Merge default OBJ scripts with model-specific overrides.
        
        Base defaults are selected by gender + traffic type.
        Model config can override any specific protocol.
        """
        from obj_defaults import DEFAULT_OBJ_FEMALE, DEFAULT_OBJ_MALE, DEFAULT_OBJ_MALE_GAY
        
        gk = self._resolve_gender_key()
        
        # Select base template
        if gk == "male_gay":
            base = DEFAULT_OBJ_MALE_GAY
        elif gk == "male":
            base = DEFAULT_OBJ_MALE
        else:
            base = DEFAULT_OBJ_FEMALE
        
        # Merge: base defaults + model-specific overrides
        model_overrides = self.c.get("obj_scripts", {})
        if model_overrides:
            # Model has custom scripts — use base as fallback for missing protocols
            merged = {**base, **model_overrides}
        else:
            # No custom scripts — use full base
            merged = dict(base)
        
        self.c["obj_scripts"] = merged

    def _build_sexting_sequences(self):
        """Build personalized sexting sequences from templates.
        
        Returns dict: {sheet_name: [(name, text, note, msg_type), ...]}
        Uses model's voice, pet names, and emoji preferences.
        """
        from sexting_defaults import SEXTING_TEMPLATES, DYNAMIC_ORDER, DYNAMIC_SHEET_NAMES
        
        gk = self._resolve_gender_key()
        
        # Determine how many sequences to generate
        num_sets = self.c.get("sexting_sets", 4)  # default: 4 dynamics
        dynamics_to_gen = DYNAMIC_ORDER[:num_sets]
        
        # Model personalization tokens
        pet_names = self.c.get("voice_pet_names", "babe, baby")
        pets = [p.strip() for p in pet_names.split(",") if p.strip()]
        pet1 = pets[0] if pets else "babe"
        pet2 = pets[1] if len(pets) > 1 else pet1
        
        # Pick emoji based on gender
        if gk == "female":
            emoji = "🖤"
        else:
            emoji = "💪"
        
        # Non-explicit models: skip sexting sequences entirely
        if self.c.get("explicit_level") in ("non_explicit",):
            return {}
        
        result = {}
        for dynamic in dynamics_to_gen:
            key = (gk, dynamic)
            template = SEXTING_TEMPLATES.get(key)
            if not template:
                continue
            
            sheet_name = DYNAMIC_SHEET_NAMES[dynamic]
            personalized = []
            for name, text, note, msg_type in template:
                # Replace placeholders
                p_text = text.replace("{pet}", pet1).replace("{pet2}", pet2).replace("{emoji}", emoji)
                p_note = note.replace("{pet}", pet1).replace("{pet2}", pet2).replace("{emoji}", emoji) if note else None
                personalized.append((name, p_text, p_note, msg_type))
            
            result[sheet_name] = personalized
        
        return result

    def ensure_dirs(self):
        os.makedirs(self.web_dir, exist_ok=True)
        os.makedirs(self.scripts_dir, exist_ok=True)
        print(f"[DIRS] {self.web_dir}, {self.scripts_dir}")
    
    def generate_journey_xlsx(self):
        """Generate the journey workbook (returns wb, not saved yet)."""
        wb = openpyxl.Workbook()
        
        # Sheet 1: Main Journey
        sheet_name = f"{self.c['name']}Journey"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws1 = wb.active
        ws1.title = sheet_name
        setup_journey_sheet(ws1)
        add_rows_reversed(ws1, self.c["journey"])
        
        # Meetup Redirect (dating app only)
        if self.c.get("meetup_redirect"):
            ws = wb.create_sheet("MeetupRedirect")
            setup_journey_sheet(ws)
            add_rows_reversed(ws, self.c["meetup_redirect"])
        
        # NR Waves
        if self.c.get("nr_waves"):
            ws = wb.create_sheet("NRWaves")
            setup_journey_sheet(ws)
            add_rows_reversed(ws, self.c["nr_waves"])
        
        # Personal Info
        if self.c.get("personal_info"):
            ws = wb.create_sheet(f"Personal{self.c['name']}"[:31])
            setup_journey_sheet(ws)
            add_standalone_rows(ws, self.c["personal_info"])
        
        # Positive Spin
        if self.c.get("positive_spin"):
            ws = wb.create_sheet("PositiveSpin")
            setup_journey_sheet(ws)
            add_standalone_rows(ws, self.c["positive_spin"])
        
        # Location Handling (dating app only)
        if self.c.get("location_handling"):
            ws = wb.create_sheet("LocationHandling")
            setup_journey_sheet(ws)
            add_standalone_rows(ws, self.c["location_handling"])
        
        # Re-Engagement
        if self.c.get("re_engagement"):
            ws = wb.create_sheet("ReEngagement")
            setup_journey_sheet(ws)
            add_rows_reversed(ws, self.c["re_engagement"])
        
        # Standalone Sexting Sequences (sex2, sex3, sex4, sex5)
        sexting_seqs = self._build_sexting_sequences()
        for sheet_name, rows in sexting_seqs.items():
            ws = wb.create_sheet(sheet_name)
            setup_journey_sheet(ws)
            # Convert 4-tuples to 3-tuples (drop msg_type for XLSX, keep for styling)
            xlsx_rows = []
            for name, text, note, msg_type in rows:
                # Prefix sheet name to Name for uniqueness across all sheets
                unique_name = f"{sheet_name} {name}" if not name.startswith(sheet_name) else name
                if len(unique_name) > 64:
                    unique_name = unique_name[:64]
                xlsx_rows.append((unique_name, text, note, msg_type))
            # Use add_rows_reversed with 4-tuples (it only uses first 3)
            add_rows_reversed(ws, xlsx_rows)
        
        return wb
    
    def generate_obj_xlsx(self):
        """Generate OBJ/RES/SIT workbook (returns wb)."""
        wb = openpyxl.Workbook()
        
        obj = self.c.get("obj_scripts", {})
        for sheet_name, (rows, category) in obj.items():
            fill = {"obj": OBJ_FILL, "res": RES_FILL, "sit": SIT_FILL}.get(category, OBJ_FILL)
            make_obj_sheet(wb, sheet_name, rows, fill)
        
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        return wb
    
    def generate_merged_xlsx(self):
        """Generate and save the merged XLSX."""
        wb_j = self.generate_journey_xlsx()
        wb_o = self.generate_obj_xlsx()
        
        clean_name = self.c['name'].strip().replace(" ", "_")
        output = os.path.join(self.web_dir, f"{clean_name}_Complete_Infloww.xlsx")
        n_sheets, n_scripts = merge_workbooks(wb_j, wb_o, output)
        print(f"[XLSX] {output} — {n_sheets} sheets, {n_scripts} scripts")
        return output
    
    def download_profile_photo(self):
        """Download profile photo from Airtable."""
        photo = download_photo(self.c["airtable_name"], self.web_dir)
        if photo:
            self.c["photo_file"] = photo
        return photo

    # ══════════════════════════════════════════
    # HTML GENERATORS
    # ══════════════════════════════════════════

    def _build_journey_phases(self):
        """Group journey messages into collapsible display phases."""
        phases = []
        current = None
        sext_count = 0
        after_ppv = False
        PHASE_META = {
            "rapport": ("Rapport", "#3fb950", "\U0001f4ac"),
            "teasing": ("Teasing", "#bc8cff", "\U0001f525"),
            "aftercare": ("Aftercare", "#58a6ff", "\U0001f4aa"),
        }
        for msg_id, text, note, msg_type in self.c.get("journey", []):
            if msg_type == "ppv":
                # PPV stays in its current phase (e.g., PPV 0 stays in Teasing)
                # Only defaults to "sext" if there's no current phase
                cat = current["cat"] if current else "sext"
            elif msg_type == "wait":
                cat = current["cat"] if current else "sext"
            else:
                cat = msg_type
            need_new = (current is None or cat != current["cat"]
                        or (cat == "sext" and msg_type == "sext" and after_ppv))
            if need_new:
                if cat == "sext":
                    sext_count += 1
                    # Non-explicit models use "Intimate Phase" instead of "Sexting Phase"
                    phase_label = "Intimate Phase" if self.c.get("explicit_level") in ("non_explicit", "soft") else "Sexting Phase"
                    meta = ("%s %d" % (phase_label, sext_count), "#f85149", "\U0001f525")
                else:
                    meta = PHASE_META.get(cat, (cat.title(), "#8b949e", "\U0001f4ac"))
                current = {
                    "cat": cat, "name": meta[0], "color": meta[1],
                    "emoji": meta[2], "msgs": [], "id": "phase-%d" % len(phases),
                }
                phases.append(current)
            current["msgs"].append((msg_id, text, note, msg_type))
            after_ppv = (msg_type == "ppv")
        return phases

    # Intensity targets per sexting message ID
    _INTENSITY = {
        "S1-1": 6, "S1-2": 7, "S1-3": 8, "S1-5": 8,
        "S1-6": 6, "S1-7": 8, "S1-8": 8, "S1-9": 9, "S1-11": 9,
        "S1-12": 9, "S1-13": 9, "S1-14": 9, "S1-15": 10, "S1-17": 10,
        "S1-18": 10, "S1-19": 10, "S1-20": 10, "S1-22": 10,
    }
    _INTENSITY_NE = {
        "S1-1": 4, "S1-2": 5, "S1-3": 6, "S1-5": 6,
        "S1-6": 5, "S1-7": 6, "S1-8": 6, "S1-9": 7, "S1-11": 7,
        "S1-12": 7, "S1-13": 7, "S1-14": 7, "S1-15": 8, "S1-17": 8,
        "S1-18": 8, "S1-19": 8, "S1-20": 8, "S1-22": 8,
    }

    @staticmethod
    def _intensity_html(level):
        """Return HTML for a compact intensity pill badge."""
        if level <= 0:
            return ''
        if level <= 4:
            bg = '#3b82f620'; border = '#3b82f650'; color = '#60a5fa'   # blue
        elif level <= 6:
            bg = '#eab30820'; border = '#eab30850'; color = '#facc15'   # yellow
        elif level <= 8:
            bg = '#f9731620'; border = '#f9731650'; color = '#fb923c'   # orange
        else:
            bg = '#ef444420'; border = '#ef444450'; color = '#f87171'   # red
        return (
            '<span class="intensity" title="Intensity %d/10" style="background:%s;border-color:%s;color:%s">'
            '%d</span>'
            % (level, bg, border, color, level)
        )

    def generate_guide_html(self):
        """Generate index.html \u2014 main chatter guide page."""
        h = html_mod.escape
        c = self.c
        name = c["name"]
        name_up = name.upper()
        photo = h(c.get("photo_file", "profile.jpg"))
        phases = self._build_journey_phases()
        is_ne = c.get("explicit_level") in ("non_explicit", "soft")
        imap = self._INTENSITY_NE if is_ne else self._INTENSITY

        # \u2500\u2500 Journey HTML \u2500\u2500
        journey_parts = []
        toc_links = []
        for i, phase in enumerate(phases):
            pid = phase["id"]
            tcls = ' class="toc-ppv"' if phase["cat"] == "sext" else ''
            toc_links.append('<a href="#%s"%s>%s %s</a>' % (pid, tcls, phase["emoji"], h(phase["name"])))
            msg_count = sum(1 for m in phase["msgs"] if m[3] != "ppv")
            ppv_count = sum(1 for m in phase["msgs"] if m[3] == "ppv")
            rows = []
            for mid, txt, note, mtype in phase["msgs"]:
                et = h(txt)
                ilevel = imap.get(mid, 0)
                ibar = self._intensity_html(ilevel)
                if mtype == "ppv":
                    rows.append(
                        '<div class="ppv-moment"><span class="ppv-icon">\U0001f4ce</span>'
                        '<span class="ppv-label">%s</span>%s</div>' % (et, ibar))
                else:
                    wcls = ' msg-wait' if mtype == 'wait' else ''
                    note_h = '<div class="msg-note">%s</div>' % h(note) if note else ''
                    rows.append(
                        '<div class="msg%s"><span class="msg-id">%s</span>'
                        '<div class="msg-body"><div class="msg-text">%s</div>%s</div>'
                        '%s'
                        '<button class="cp" data-copy="%s">'
                        '<span class="cp-icon">\U0001f4cb</span></button></div>'
                        % (wcls, h(mid), et, note_h, ibar, et))
            collapsed = ' collapsed' if i > 0 else ''
            # Build smart meta label: "X msgs" or "PPV" or "X msgs + PPV"
            if msg_count > 0 and ppv_count > 0:
                meta_label = "%d msgs + %d PPV" % (msg_count, ppv_count)
            elif msg_count > 0:
                meta_label = "%d msgs" % msg_count
            elif ppv_count > 0:
                meta_label = "%d PPV" % ppv_count
            else:
                meta_label = "0 msgs"
            # Collect all copyable texts for Copy All button
            all_texts = [txt for mid, txt, note, mtype in phase["msgs"] if mtype not in ("ppv", "wait")]
            copy_all_data = '\\n\\n'.join(t.replace('"', '&quot;') for t in all_texts)
            copy_all_btn = ('<button class="copy-all" data-copy-all="%s">'
                           'Copy All</button>' % copy_all_data) if all_texts else ''
            journey_parts.append(
                '<div class="phase%s" id="%s" style="--phase-color:%s">\n'
                '  <div class="phase-header">\n'
                '    <span class="phase-emoji">%s</span>\n'
                '    <span class="phase-name">%s</span>\n'
                '    <span class="phase-meta">%s</span>\n'
                '    %s\n'
                '    <span class="toggle">\u25bc</span>\n'
                '  </div>\n'
                '  <div class="phase-body">\n%s\n  </div>\n</div>'
                % (collapsed, pid, phase["color"], phase["emoji"],
                   h(phase["name"]), meta_label, copy_all_btn, '\n'.join(rows)))

        # \u2500\u2500 NR Waves \u2500\u2500
        nr_html = ''
        if c.get('nr_waves'):
            nr_rows = []
            for mid, txt, note, mtype in c['nr_waves']:
                nr_rows.append(
                    '<tr><td><strong>%s</strong></td><td>%s</td>'
                    '<td class="nr-note">%s</td>'
                    '<td><button class="cp" data-copy="%s">'
                    '<span class="cp-icon">\U0001f4cb</span></button></td></tr>'
                    % (h(mid), h(txt), h(note) if note else '', h(txt)))
            nr_html = (
                '<div class="ref-section" id="nr-waves">\n'
                '  <div class="ref-title">\U0001f507 NR Wave Protocol</div>\n'
                '  <table class="nr-table"><thead><tr>'
                '<th>Wave</th><th>Message</th><th>Note</th><th></th>'
                '</tr></thead>\n  <tbody>'
                + ''.join(nr_rows)
                + '</tbody></table>\n</div>\n<hr class="divider">\n')

        # \u2500\u2500 Personal Info \u2500\u2500
        pi_html = ''
        if c.get('personal_info'):
            items = []
            for label, txt, note in c['personal_info']:
                items.append(
                    '<div class="qi-item"><span class="qi-label">%s</span>'
                    '<span class="qi-text">%s</span>'
                    '<button class="cp" data-copy="%s">'
                    '<span class="cp-icon">\U0001f4cb</span></button></div>'
                    % (h(label), h(txt), h(txt)))
            pi_html = (
                '<div class="ref-section" id="personal-info">\n'
                '  <div class="ref-title">\U0001f464 Personal Info</div>\n'
                '  <div class="qi-grid">' + ''.join(items) + '</div>\n'
                '</div>\n<hr class="divider">\n')

        # \u2500\u2500 Positive Spin \u2500\u2500
        spin_html = ''
        if c.get('positive_spin'):
            items = []
            for label, txt, note in c['positive_spin']:
                items.append(
                    '<div class="qi-item"><span class="qi-label">%s</span>'
                    '<span class="qi-text">%s</span>'
                    '<button class="cp" data-copy="%s">'
                    '<span class="cp-icon">\U0001f4cb</span></button></div>'
                    % (h(label), h(txt), h(txt)))
            spin_html = (
                '<div class="ref-section" id="positive-spin">\n'
                '  <div class="ref-title">\u2728 Positive Spin</div>\n'
                '  <div class="qi-grid">' + ''.join(items) + '</div>\n'
                '</div>\n<hr class="divider">\n')

        # \u2500\u2500 Meetup Redirect (dating_app only) \u2500\u2500
        mr_html = ''
        if c.get('traffic') == 'dating_app' and c.get('meetup_redirect'):
            mr_rows = []
            for mid, txt, note, mtype in c['meetup_redirect']:
                et = h(txt)
                note_h = '<div class="msg-note">%s</div>' % h(note) if note else ''
                mr_rows.append(
                    '<div class="msg"><span class="msg-id">%s</span>'
                    '<div class="msg-body"><div class="msg-text">%s</div>%s</div>'
                    '<button class="cp" data-copy="%s">'
                    '<span class="cp-icon">\U0001f4cb</span></button></div>'
                    % (h(mid), et, note_h, et))
            mr_html = (
                '<div class="phase collapsed" id="meetup-redirect" style="--phase-color:#f85149">\n'
                '  <div class="phase-header">\n'
                '    <span class="phase-emoji">\U0001f91d</span>\n'
                '    <span class="phase-name">Meetup Redirect</span>\n'
                '    <span class="phase-meta">%d msgs</span>\n'
                '    <span class="toggle">\u25bc</span>\n'
                '  </div>\n'
                '  <div class="phase-body">\n%s\n  </div>\n</div>\n'
                % (len(mr_rows), '\n'.join(mr_rows)))

        # \u2500\u2500 Location Handling \u2500\u2500
        loc_html = ''
        if c.get('location_handling'):
            items = []
            for label, txt, note in c['location_handling']:
                items.append(
                    '<div class="qi-item"><span class="qi-label">%s</span>'
                    '<span class="qi-text">%s</span>'
                    '<button class="cp" data-copy="%s">'
                    '<span class="cp-icon">\U0001f4cb</span></button></div>'
                    % (h(label), h(txt), h(txt)))
            loc_html = (
                '<div class="ref-section" id="location">\n'
                '  <div class="ref-title">\U0001f4cd Location Handling</div>\n'
                '  <div class="qi-grid">' + ''.join(items) + '</div>\n'
                '</div>\n<hr class="divider">\n')

        # \u2500\u2500 Re-engagement \u2500\u2500
        re_html = ''
        if c.get('re_engagement'):
            re_rows = []
            for mid, txt, note, mtype in c['re_engagement']:
                et = h(txt)
                note_h = '<div class="msg-note">%s</div>' % h(note) if note else ''
                re_rows.append(
                    '<div class="msg"><span class="msg-id">%s</span>'
                    '<div class="msg-body"><div class="msg-text">%s</div>%s</div>'
                    '<button class="cp" data-copy="%s">'
                    '<span class="cp-icon">\U0001f4cb</span></button></div>'
                    % (h(mid), et, note_h, et))
            re_html = (
                '<div class="phase collapsed" id="re-engagement" style="--phase-color:#58a6ff">\n'
                '  <div class="phase-header">\n'
                '    <span class="phase-emoji">\U0001f504</span>\n'
                '    <span class="phase-name">Re-engagement</span>\n'
                '    <span class="phase-meta">%d msgs</span>\n'
                '    <span class="toggle">\u25bc</span>\n'
                '  </div>\n'
                '  <div class="phase-body">\n%s\n  </div>\n</div>\n'
                % (len(re_rows), '\n'.join(re_rows)))

        # \u2500\u2500 Profile grid \u2500\u2500
        interests = c.get('interests', [])
        interests_str = ', '.join(interests) if isinstance(interests, list) else str(interests)
        info_pairs = [
            ('Nationality', c.get('nationality', '')),
            ('Page Type', c.get('page_type', '')),
            ('Traffic', c.get('traffic', '')),
            ('Job', c.get('job', '')),
            ('Interests', interests_str),
            ('Physical', c.get('physical', '')),
            ('Countries', c.get('countries', '')),
            ('Languages', c.get('languages', '')),
        ]
        profile_grid = ''.join(
            '<div class="profile-item"><div class="label">%s</div>'
            '<div class="value">%s</div></div>' % (h(lbl), h(val))
            for lbl, val in info_pairs if val)

        # \u2500\u2500 Voice \u2500\u2500
        voice_html = ''
        if c.get('voice'):
            vp = '<strong>\U0001f5e3\ufe0f Voice:</strong> %s' % h(c['voice'])
            if c.get('voice_pet_names'):
                vp += '<br><strong>Pet names:</strong> %s' % h(c['voice_pet_names'])
            if c.get('voice_never'):
                vp += '<br><strong>Never say:</strong> %s' % h(c['voice_never'])
            voice_html = '    <div class="voice-box">%s</div>\n' % vp

        # \u2500\u2500 TOC reference links \u2500\u2500
        ref_toc = []
        if c.get('nr_waves'):
            ref_toc.append('<a href="#nr-waves" class="toc-ref">NR</a>')
        if c.get('traffic') == 'dating_app' and c.get('location_handling'):
            ref_toc.append('<a href="#location" class="toc-ref">Location</a>')
        if c.get('personal_info'):
            ref_toc.append('<a href="#personal-info" class="toc-ref">Personal</a>')
        if c.get('positive_spin'):
            ref_toc.append('<a href="#positive-spin" class="toc-ref">Spin</a>')
        ref_toc.append(
            '<a href="objections.html" class="toc-ref"'
            ' style="color:var(--red);border-color:#f8514933">'
            '\U0001f6e1\ufe0f OBJ/RES</a>')
        toc_str = '\n  '.join(toc_links + ref_toc)

        # \u2500\u2500 CSS \u2500\u2500
        CSS = (
            '<style>\n'
            '*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}\n'
            ':root{--bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--muted:#8b949e;'
            '--accent:#58a6ff;--green:#3fb950;--red:#f85149;--orange:#d29922;--purple:#bc8cff;--yellow:#e3b341}\n'
            'html{scroll-behavior:smooth}\n'
            'body{background:var(--bg);color:var(--text);font-family:"Segoe UI",system-ui,sans-serif;line-height:1.6;padding:0 0 60px}\n'
            '.page{max-width:960px;margin:0 auto;padding:16px 20px}\n'
            '.back-btn{display:inline-flex;align-items:center;gap:6px;color:var(--muted);text-decoration:none;font-size:.85rem;padding:10px 0;transition:.15s}\n'
            '.back-btn:hover{color:var(--accent)}\n'
            '.hero{display:flex;align-items:center;gap:18px;padding:20px 0 16px;border-bottom:1px solid var(--border)}\n'
            '.hero-photo{width:72px;height:72px;border-radius:50%;object-fit:cover;border:3px solid var(--accent);flex-shrink:0}\n'
            '.hero-info{flex:1}.hero h1{font-size:1.8rem;font-weight:800;letter-spacing:2px;line-height:1.2}\n'
            '.hero .tagline{color:var(--muted);font-size:.85rem}\n'
            '.toc{position:sticky;top:0;z-index:50;background:var(--bg);border-bottom:1px solid var(--border);'
            'padding:8px 0;margin:0 -20px;padding-left:20px;padding-right:20px;display:flex;gap:6px;'
            'overflow-x:auto;-webkit-overflow-scrolling:touch}\n'
            '.toc.shadow{box-shadow:0 4px 16px rgba(0,0,0,.4)}\n'
            '.toc a{font-size:.72rem;color:var(--muted);background:var(--card);border:1px solid var(--border);'
            'padding:4px 10px;border-radius:14px;text-decoration:none;white-space:nowrap;transition:.15s;flex-shrink:0}\n'
            '.toc a:hover,.toc a.active{color:var(--accent);border-color:var(--accent)}\n'
            '.toc a.toc-ppv{color:var(--green);border-color:#3fb95033}\n'
            '.toc a.toc-ref{color:var(--orange);border-color:#d2992233}\n'
            '.info-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:20px 0}\n'
            '.profile{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px 18px}\n'
            '.profile h2{font-size:.78rem;color:var(--accent);text-transform:uppercase;letter-spacing:1.5px;margin-bottom:12px}\n'
            '.profile-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px 16px}\n'
            '.profile-item .label{color:var(--muted);font-size:.7rem;text-transform:uppercase;letter-spacing:.5px}\n'
            '.profile-item .value{font-size:.85rem}\n'
            '.voice-box{background:#1c262f;border-left:3px solid var(--purple);padding:10px 14px;'
            'border-radius:0 8px 8px 0;margin-top:12px;font-size:.84rem;color:#c4b5e0}\n'
            '.voice-box strong{color:var(--purple)}\n'
            '.obj-link{display:flex;align-items:center;gap:14px;background:#f8514910;border:1px solid #f8514933;'
            'border-radius:12px;padding:14px 18px;margin:16px 0;text-decoration:none;transition:.15s}\n'
            '.obj-link:hover{border-color:var(--red)}\n'
            '.phase{margin:10px 0 0;border:1px solid var(--border);border-radius:12px;overflow:hidden}\n'
            '.phase-header{display:flex;align-items:center;gap:10px;padding:12px 16px;cursor:pointer;'
            'user-select:none;background:var(--card);border-left:3px solid var(--phase-color,var(--accent))}\n'
            '.phase-header:hover{background:#1c2129}\n'
            '.phase-header .phase-emoji{font-size:1.2rem}\n'
            '.phase-header .phase-name{font-size:1rem;font-weight:700;flex:1}\n'
            '.phase-header .phase-meta{font-size:.72rem;color:var(--muted);background:var(--bg);'
            'padding:2px 10px;border-radius:12px;border:1px solid var(--border)}\n'
            '.phase-header .toggle{color:var(--muted);font-size:.8rem;transition:transform .2s}\n'
            '.phase.collapsed .toggle{transform:rotate(-90deg)}\n'
            '.phase-body{padding:4px 0 8px}.phase.collapsed .phase-body{display:none}\n'
            '.msg{display:flex;gap:10px;align-items:flex-start;padding:8px 16px;margin:2px 0;transition:background .12s}\n'
            '.msg:hover{background:#161b2266}\n'
            '.msg-id{font-size:.64rem;color:var(--muted);background:var(--card);border:1px solid var(--border);'
            'padding:2px 6px;border-radius:4px;white-space:nowrap;margin-top:4px;flex-shrink:0;'
            'min-width:36px;text-align:center;font-weight:600}\n'
            '.msg-body{flex:1;min-width:0}\n'
            '.msg-text{font-size:1rem;line-height:1.5;word-wrap:break-word}\n'
            '.msg-note{background:#e3b34112;border-left:3px solid var(--yellow);padding:5px 10px;'
            'border-radius:0 6px 6px 0;margin:4px 0 2px;font-size:.8rem;color:var(--yellow)}\n'
            '.msg-wait{background:#d2992210}.msg-wait .msg-id{border-color:var(--orange);color:var(--orange)}\n'
            '.intensity{display:inline-flex;align-items:center;justify-content:center;'
            'min-width:20px;height:20px;padding:0 5px;border-radius:10px;'
            'font-size:.62rem;font-weight:700;border:1px solid;flex-shrink:0;'
            'letter-spacing:-.02em;margin-left:4px}\n'
            '.ppv-moment{background:#3fb95010;border:1px solid #3fb95028;border-radius:10px;'
            'padding:8px 14px;margin:6px 16px;display:flex;align-items:center;gap:10px}\n'
            '.ppv-moment .ppv-icon{font-size:1.2rem}\n'
            '.ppv-moment .ppv-label{font-size:.88rem;font-weight:600;color:var(--green)}\n'
            '.cp{background:none;border:1px solid transparent;color:var(--border);width:28px;height:28px;'
            'border-radius:6px;cursor:pointer;font-size:.75rem;flex-shrink:0;display:flex;align-items:center;'
            'justify-content:center;transition:.15s;position:relative;margin-top:2px;opacity:.3}\n'
            '.msg:hover .cp,.qi-item:hover .cp,.nr-table tr:hover .cp{opacity:1;border-color:var(--border)}\n'
            '.cp:hover{background:var(--card);color:var(--text);opacity:1}\n'
            '.cp.ok{border-color:var(--green);color:var(--green);opacity:1}\n'
            '.cp.ok::after{content:"\\2713";position:absolute;font-size:.7rem;font-weight:700}\n'
            '.cp.ok .cp-icon{visibility:hidden}\n'
            '.divider{border:none;border-top:1px solid var(--border);margin:24px 0}\n'
            '.ref-section{margin:16px 0}\n'
            '.ref-title{font-size:.95rem;font-weight:700;padding:8px 0;border-bottom:2px solid var(--border);margin-bottom:8px}\n'
            '.nr-table{width:100%;border-collapse:collapse;margin:6px 0}\n'
            '.nr-table th{text-align:left;font-size:.68rem;color:var(--muted);text-transform:uppercase;'
            'padding:5px 10px;border-bottom:1px solid var(--border)}\n'
            '.nr-table td{padding:7px 10px;font-size:.88rem;border-bottom:1px solid #21262d;vertical-align:top}\n'
            '.nr-note{color:var(--muted);font-size:.8rem}\n'
            '.qi-grid{display:grid;grid-template-columns:1fr 1fr;gap:5px;margin:6px 0}\n'
            '.qi-item{display:flex;gap:8px;align-items:flex-start;padding:8px 12px;background:var(--card);'
            'border:1px solid var(--border);border-radius:8px}\n'
            '.qi-label{font-size:.7rem;font-weight:600;color:var(--accent);text-transform:uppercase;'
            'min-width:56px;margin-top:3px;flex-shrink:0}\n'
            '.qi-text{flex:1;font-size:.88rem}\n'
            '.top-btn{position:fixed;bottom:16px;right:16px;width:36px;height:36px;border-radius:50%;'
            'background:var(--card);border:1px solid var(--border);color:var(--muted);font-size:1rem;'
            'cursor:pointer;display:none;align-items:center;justify-content:center;z-index:100;transition:.15s}\n'
            '.top-btn:hover{color:var(--text);border-color:var(--muted)}.top-btn.show{display:flex}\n'
            # Search box styles
            '.search-wrap{margin:12px 0 4px;position:relative}\n'
            '.search-input{width:100%;background:var(--card);border:1px solid var(--border);color:var(--text);'
            'padding:8px 14px 8px 36px;border-radius:8px;font-size:.85rem;outline:none;transition:.15s}\n'
            '.search-input:focus{border-color:var(--accent);box-shadow:0 0 0 2px #58a6ff22}\n'
            '.search-input::placeholder{color:var(--muted)}\n'
            '.search-icon{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:var(--muted);font-size:.8rem;pointer-events:none}\n'
            '.search-count{position:absolute;right:12px;top:50%;transform:translateY(-50%);color:var(--muted);font-size:.72rem}\n'
            '.msg.search-hide,.phase.search-hide{display:none}\n'
            '.msg.search-match .msg-text{background:#58a6ff12;border-radius:4px}\n'
            # Copy-all button styles
            '.copy-all{background:none;border:1px solid var(--border);color:var(--muted);padding:2px 8px;'
            'border-radius:6px;font-size:.64rem;cursor:pointer;transition:.15s;margin-left:auto;margin-right:8px}\n'
            '.copy-all:hover{border-color:var(--accent);color:var(--accent)}\n'
            '.copy-all.ok{border-color:var(--green);color:var(--green)}\n'
            '@media(max-width:700px){.page{padding:12px 10px}.info-grid,.qi-grid{grid-template-columns:1fr}'
            '.hero{flex-direction:column;text-align:center}.hero h1{font-size:1.5rem}'
            '.profile-grid{grid-template-columns:1fr}.toc{margin:0 -10px;padding-left:10px;padding-right:10px}}\n'
            '</style>')

        # \u2500\u2500 JS \u2500\u2500
        JS = (
            '<script>\n'
            # Copy individual message
            "document.addEventListener('click',e=>{"
            "const b=e.target.closest('.cp');if(b){"
            "navigator.clipboard.writeText(b.dataset.copy).then(()=>{b.classList.add('ok');"
            "setTimeout(()=>b.classList.remove('ok'),1200)});return;}"
            # Copy All button
            "const ca=e.target.closest('.copy-all');if(ca){"
            "e.stopPropagation();"
            "const txt=ca.dataset.copyAll.replace(/\\\\n/g,'\\n');"
            "navigator.clipboard.writeText(txt).then(()=>{ca.classList.add('ok');"
            "ca.textContent='Copied!';"
            "setTimeout(()=>{ca.classList.remove('ok');ca.textContent='Copy All'},1500)});"
            "return;}});\n"
            # Collapse/expand phases
            "document.querySelectorAll('.phase-header').forEach(h=>{h.addEventListener('click',e=>{"
            "if(e.target.closest('.cp')||e.target.closest('.copy-all'))return;"
            "h.closest('.phase').classList.toggle('collapsed')})});\n"
            # TOC expand on click
            'document.querySelectorAll(\'.toc a[href^="#"]\').forEach(a=>{a.addEventListener(\'click\',()=>{'
            "const t=document.getElementById(a.getAttribute('href').slice(1));"
            "if(t&&t.classList.contains('phase'))t.classList.remove('collapsed')})});\n"
            # Scroll to top
            "const topBtn=document.getElementById('top-btn');\n"
            "window.addEventListener('scroll',()=>topBtn.classList.toggle('show',window.scrollY>400));\n"
            "topBtn.addEventListener('click',()=>window.scrollTo({top:0,behavior:'smooth'}));\n"
            # TOC shadow
            "const toc=document.getElementById('toc');\n"
            "window.addEventListener('scroll',()=>toc.classList.toggle('shadow',window.scrollY>200));\n"
            # Active TOC tracking
            'const tocLinks=document.querySelectorAll(\'.toc a[href^="#"]\');\n'
            "const sections=Array.from(tocLinks).map(a=>({link:a,"
            "el:document.getElementById(a.getAttribute('href').slice(1))})).filter(s=>s.el);\n"
            "function updateToc(){let a=sections[0];for(const s of sections)"
            "if(s.el.getBoundingClientRect().top<=100)a=s;"
            "tocLinks.forEach(l=>l.classList.remove('active'));if(a)a.link.classList.add('active')}\n"
            "window.addEventListener('scroll',updateToc);updateToc();\n"
            # Search functionality
            "const searchInput=document.getElementById('search');\n"
            "const searchCount=document.getElementById('search-count');\n"
            "searchInput.addEventListener('input',()=>{"
            "const q=searchInput.value.toLowerCase().trim();"
            "const msgs=document.querySelectorAll('.msg');"
            "const phases=document.querySelectorAll('.phase');"
            "if(!q){msgs.forEach(m=>{m.classList.remove('search-hide','search-match')});"
            "phases.forEach(p=>p.classList.remove('search-hide'));searchCount.textContent='';return;}"
            "let count=0;"
            "phases.forEach(p=>{"
            "let hasMatch=false;"
            "p.querySelectorAll('.msg').forEach(m=>{"
            "const txt=m.querySelector('.msg-text');"
            "if(txt&&txt.textContent.toLowerCase().includes(q)){"
            "m.classList.remove('search-hide');m.classList.add('search-match');"
            "hasMatch=true;count++;"
            "}else{m.classList.add('search-hide');m.classList.remove('search-match')}});"
            "if(hasMatch){p.classList.remove('search-hide','collapsed')}"
            "else{p.classList.add('search-hide')}});"
            "searchCount.textContent=count+' found'});\n"
            # Keyboard shortcut: Escape to clear search
            "searchInput.addEventListener('keydown',e=>{"
            "if(e.key==='Escape'){searchInput.value='';searchInput.dispatchEvent(new Event('input'));"
            "searchInput.blur()}});\n"
            '</script>')

        # \u2500\u2500 Assemble page \u2500\u2500
        page = (
            '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
            '<meta charset="UTF-8">\n'
            '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
            '<title>%s \u2014 Chatter Guide</title>\n' % h(name_up)
            + CSS + '\n</head>\n<body>\n<div class="page">\n\n'
            '<a href="../" class="back-btn">\u2190 All Creators</a>\n\n'
            '<div class="hero" id="top">\n'
            '  <img src="%s" alt="%s" class="hero-photo">\n'
            '  <div class="hero-info">\n    <h1>%s</h1>\n'
            '    <div class="tagline">%s \u00b7 %s \u00b7 %s</div>\n'
            '  </div>\n</div>\n\n'
            % (photo, h(name), h(name_up),
               c.get('age', ''), h(c.get('location', '')), h(c.get('job', '')))
            + '<nav class="toc" id="toc">\n  %s\n</nav>\n\n' % toc_str
            + '<div class="search-wrap">\n'
            '  <span class="search-icon">&#128269;</span>\n'
            '  <input type="text" class="search-input" id="search" '
            'placeholder="Search scripts..." autocomplete="off">\n'
            '  <span class="search-count" id="search-count"></span>\n'
            '</div>\n\n'
            + '<div class="info-grid">\n  <div>\n'
            '    <div class="profile">\n'
            '      <h2>\U0001f464 Character Profile</h2>\n'
            '      <div class="profile-grid">%s</div>\n'
            '    </div>\n' % profile_grid
            + voice_html
            + '  </div>\n')
        # Right column: special notes, key phrases, explicit level
        right_content = ''
        if c.get('special_notes') or c.get('key_phrases') or c.get('explicit_level'):
            right_parts = []
            if c.get('explicit_level'):
                level_map = {'full': ('Full Explicit', '#f85149', '#f8514912', '#f8514933'),
                             'non_explicit': ('Non-Explicit (Lingerie/Tease)', '#d29922', '#d2992212', '#d2992233'),
                             'soft': ('Soft (No Hardcore)', '#e3b341', '#e3b34112', '#e3b34133')}
                defaults = (c['explicit_level'], '#8b949e', '#8b949e12', '#8b949e33')
                lbl, clr, bg, bdr = level_map.get(c['explicit_level'], defaults)
                right_parts.append(
                    '<div class="content-level" style="background:%s;border:1px solid %s;'
                    'padding:8px 14px;border-radius:8px;margin-bottom:10px;font-size:.84rem">'
                    '<strong style="color:%s">Content Level:</strong> %s</div>'
                    % (bg, bdr, clr, h(lbl)))
            if c.get('special_notes'):
                right_parts.append(
                    '<div style="background:#e3b34112;border-left:3px solid var(--yellow);'
                    'padding:10px 14px;border-radius:0 8px 8px 0;margin-bottom:10px;'
                    'font-size:.84rem;color:var(--yellow)"><strong>Notes:</strong> %s</div>'
                    % h(c['special_notes']))
            if c.get('key_phrases'):
                phrases = ''.join('<span style="background:var(--card);border:1px solid var(--border);'
                                  'padding:3px 8px;border-radius:6px;font-size:.8rem;display:inline-block;'
                                  'margin:2px 3px">%s</span>' % h(p) for p in c['key_phrases'][:8])
                right_parts.append(
                    '<div style="margin-top:4px"><strong style="font-size:.78rem;color:var(--muted);'
                    'text-transform:uppercase;letter-spacing:.5px">Key Phrases:</strong><div style="'
                    'margin-top:6px;line-height:2">%s</div></div>' % phrases)
            right_content = (
                '  <div>\n    <div class="profile" style="border-color:var(--border)">\n'
                '      <h2>Quick Reference</h2>\n      %s\n    </div>\n  </div>\n'
                % '\n      '.join(right_parts))
        else:
            right_content = '  <div></div>\n'
        page += right_content + '</div>\n\n'
        page += (
            '<a href="objections.html" class="obj-link">\n'
            '  <span style="font-size:1.4rem">\U0001f6e1\ufe0f</span>\n'
            '  <div style="flex:1">\n'
            '    <div style="font-size:.95rem;font-weight:700;color:var(--red)">'
            'Objection & Resistance Handling</div>\n'
            '    <div style="font-size:.78rem;color:var(--muted)">'
            'Complete sequences for Infloww</div>\n'
            '  </div>\n'
            '  <span style="color:var(--red);font-size:.85rem;font-weight:600">'
            'Open \u2192</span>\n</a>\n\n'
            + mr_html
            + '\n'.join(journey_parts) + '\n\n'
            + re_html
            + '<hr class="divider">\n\n'
            + nr_html + loc_html + pi_html + spin_html
            + '<div style="text-align:center;padding:24px 0 10px;color:var(--muted);font-size:.73rem">'
            'Chatting Wizard \u00b7 Script Manager \u00b7 2026</div>\n\n'
            '</div>\n<button class="top-btn" id="top-btn">\u2191</button>\n\n'
            + JS + '\n</body>\n</html>')

        path = os.path.join(self.web_dir, "index.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(page)
        print("[HTML] Guide: %s" % path)
        return path

    def generate_objections_html(self):
        """Generate objections.html \u2014 OBJ/RES/SIT handling page."""
        h = html_mod.escape
        c = self.c
        name = c["name"]
        name_up = name.upper()
        photo = h(c.get("photo_file", "profile.jpg"))

        obj_scripts = c.get("obj_scripts", {})

        # \u2500\u2500 Group into protocols \u2500\u2500
        protocols = {}
        for sheet_name, (rows, category) in obj_scripts.items():
            base = sheet_name.rstrip("0123456789") or sheet_name
            if base not in protocols:
                protocols[base] = {
                    "cat": category,
                    "display": base.replace("_", " ").title(),
                    "variants": [],
                }
            protocols[base]["variants"].append((sheet_name, rows))
        for p in protocols.values():
            p["variants"].sort(key=lambda x: x[0])

        categories = {
            "obj": {"label": "A. Objection Handling", "emoji": "\U0001f4b0",
                    "cls": "cat-obj", "protos": []},
            "res": {"label": "B. Resistance Handling", "emoji": "\U0001f500",
                    "cls": "cat-res", "protos": []},
            "sit": {"label": "C. Situational Scripts", "emoji": "\u26a1",
                    "cls": "cat-sit", "protos": []},
        }
        for base_name, proto in protocols.items():
            cat = proto["cat"]
            if cat in categories:
                categories[cat]["protos"].append((base_name, proto))

        total_sheets = len(obj_scripts)
        total_scripts = sum(len(rows) for rows, _ in obj_scripts.values())
        total_protos = len(protocols)

        # \u2500\u2500 TOC \u2500\u2500
        toc_cls_map = {"obj": "toc-obj", "res": "toc-res", "sit": "toc-sit"}
        toc_items = []
        for ck in ("obj", "res", "sit"):
            for base_name, proto in categories[ck]["protos"]:
                toc_items.append(
                    '<a href="#%s" class="%s">%s</a>'
                    % (h(base_name), toc_cls_map[ck], h(proto["display"])))

        # \u2500\u2500 Protocol HTML \u2500\u2500
        proto_parts = []
        first = True
        for ck in ("obj", "res", "sit"):
            cat = categories[ck]
            if not cat["protos"]:
                continue
            proto_parts.append(
                '<div class="cat-header %s"><span class="cat-emoji">%s</span>'
                '<div class="cat-info"><h2>%s</h2>'
                '<p>%d protocols</p></div></div>'
                % (cat["cls"], cat["emoji"], h(cat["label"]), len(cat["protos"])))

            for base_name, proto in cat["protos"]:
                nv = len(proto["variants"])
                max_s = max((len(rows) for _, rows in proto["variants"]), default=0)
                collapsed = '' if first else ' collapsed'
                first = False

                tabs = []
                panels = []
                for vi, (sn, rows) in enumerate(proto["variants"]):
                    act = ' active' if vi == 0 else ''
                    tabs.append(
                        '<div class="seq-tab%s" data-seq="%s">'
                        '<code>/%s</code></div>' % (act, h(sn), h(sn)))
                    steps = []
                    for si, (step_name, txt, note) in enumerate(rows):
                        et = h(txt)
                        tech = ('<div class="step-technique">%s</div>' % h(step_name)
                                if step_name else '')
                        note_h = ('<div class="step-note">%s</div>' % h(note)
                                  if note else '')
                        steps.append(
                            '<div class="step"><span class="step-num">%d</span>'
                            '<div class="step-body">%s'
                            '<div class="step-msg">%s</div>%s</div>'
                            '<button class="cp" data-copy="%s">'
                            '<span class="cp-icon">\U0001f4cb</span></button></div>'
                            % (si + 1, tech, et, note_h, et))
                    panels.append(
                        '<div class="seq-panel%s" id="%s">\n%s\n</div>'
                        % (act, h(sn), '\n'.join(steps)))

                # Protocol description
                desc = PROTOCOL_DESCRIPTIONS.get(base_name, "")
                desc_html = ('<div class="p-desc">%s</div>' % h(desc)) if desc else ''

                proto_parts.append(
                    '<div class="protocol%s" id="%s">\n'
                    '  <div class="protocol-header">'
                    '<span class="p-name">%s</span>'
                    '<span class="p-meta">%d steps \u00b7 %d seq</span>'
                    '<span class="toggle">\u25bc</span></div>\n'
                    '  %s\n'
                    '  <div class="protocol-body">\n'
                    '    <div class="seq-tabs">%s</div>\n%s\n'
                    '  </div>\n</div>'
                    % (collapsed, h(base_name), h(proto["display"]),
                       max_s, nv, desc_html, ''.join(tabs), '\n'.join(panels)))

        # \u2500\u2500 CSS \u2500\u2500
        CSS = (
            '<style>\n'
            '*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}\n'
            ':root{--bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--muted:#8b949e;'
            '--accent:#58a6ff;--green:#3fb950;--red:#f85149;--orange:#d29922;--purple:#bc8cff;--yellow:#e3b341}\n'
            'html{scroll-behavior:smooth}\n'
            'body{background:var(--bg);color:var(--text);font-family:"Segoe UI",system-ui,sans-serif;line-height:1.6;padding:0 0 60px}\n'
            '.page{max-width:960px;margin:0 auto;padding:16px 20px}\n'
            '.back-btn{display:inline-flex;align-items:center;gap:6px;color:var(--muted);text-decoration:none;'
            'font-size:.85rem;padding:10px 0;transition:.15s}\n'
            '.back-btn:hover{color:var(--accent)}\n'
            '.hero{display:flex;align-items:center;gap:18px;padding:20px 0 16px;border-bottom:1px solid var(--border)}\n'
            '.hero-photo{width:60px;height:60px;border-radius:50%;object-fit:cover;border:3px solid var(--accent);flex-shrink:0}\n'
            '.hero h1{font-size:1.5rem;font-weight:800;letter-spacing:2px}\n'
            '.hero .tagline{color:var(--muted);font-size:.82rem}\n'
            '.info-bar{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0 4px;font-size:.78rem}\n'
            '.info-pill{background:var(--card);border:1px solid var(--border);padding:4px 12px;'
            'border-radius:20px;display:flex;align-items:center;gap:5px;color:var(--muted)}\n'
            '.info-pill b{color:var(--text)}\n'
            '.toc{position:sticky;top:0;z-index:50;background:var(--bg);border-bottom:1px solid var(--border);'
            'padding:8px 0;margin:0 -20px;padding-left:20px;padding-right:20px;display:flex;gap:6px;'
            'overflow-x:auto;-webkit-overflow-scrolling:touch}\n'
            '.toc.shadow{box-shadow:0 4px 16px rgba(0,0,0,.4)}\n'
            '.toc a{font-size:.7rem;color:var(--muted);background:var(--card);border:1px solid var(--border);'
            'padding:4px 10px;border-radius:14px;text-decoration:none;white-space:nowrap;transition:.15s;flex-shrink:0}\n'
            '.toc a:hover,.toc a.active{color:var(--accent);border-color:var(--accent)}\n'
            '.toc a.toc-obj{color:var(--red);border-color:#f8514933}\n'
            '.toc a.toc-res{color:var(--purple);border-color:#bc8cff33}\n'
            '.toc a.toc-sit{color:var(--green);border-color:#3fb95033}\n'
            '.cat-header{margin:28px 0 8px;padding:14px 18px;border-radius:12px;display:flex;align-items:center;gap:12px}\n'
            '.cat-header.cat-obj{background:#f8514912;border:1px solid #f8514933}\n'
            '.cat-header.cat-res{background:#bc8cff12;border:1px solid #bc8cff33}\n'
            '.cat-header.cat-sit{background:#3fb95012;border:1px solid #3fb95033}\n'
            '.cat-header .cat-emoji{font-size:1.4rem}.cat-header .cat-info{flex:1}\n'
            '.cat-header h2{font-size:1.1rem;font-weight:700}\n'
            '.cat-header p{font-size:.8rem;color:var(--muted)}\n'
            '.protocol{background:var(--card);border:1px solid var(--border);border-radius:12px;margin:12px 0;overflow:hidden}\n'
            '.protocol-header{display:flex;align-items:center;gap:10px;padding:12px 16px;'
            'cursor:pointer;user-select:none;transition:background .12s}\n'
            '.protocol-header:hover{background:#1c2129}\n'
            '.protocol-header .p-name{font-size:.95rem;font-weight:600;flex:1}\n'
            '.protocol-header .p-meta{font-size:.72rem;color:var(--muted);background:var(--bg);'
            'padding:2px 8px;border-radius:10px;border:1px solid var(--border)}\n'
            '.protocol-header .toggle{color:var(--muted);font-size:.8rem;transition:transform .2s}\n'
            '.protocol.collapsed .toggle{transform:rotate(-90deg)}\n'
            '.p-desc{font-size:.78rem;color:var(--muted);padding:4px 16px 2px;border-top:1px solid var(--border);background:#1a1d23;line-height:1.4}\n'
            '.protocol.collapsed .p-desc{display:none}\n'
            '.protocol-body{padding:6px 0 12px}.protocol.collapsed .protocol-body{display:none}\n'
            '.seq-tabs{display:flex;gap:6px;padding:8px 16px}\n'
            '.seq-tab{font-size:.76rem;font-weight:700;padding:4px 12px;border-radius:8px;cursor:pointer;'
            'border:1px solid var(--border);color:var(--muted);background:var(--bg);transition:.15s}\n'
            '.seq-tab:hover{border-color:var(--accent);color:var(--accent)}\n'
            '.seq-tab.active{background:var(--accent);color:#0d1117;border-color:var(--accent)}\n'
            '.seq-panel{display:none}.seq-panel.active{display:block}\n'
            '.step{display:flex;gap:10px;padding:6px 16px;align-items:flex-start;margin:3px 0}\n'
            '.step-num{width:24px;height:24px;border-radius:50%;display:flex;align-items:center;'
            'justify-content:center;font-size:.68rem;font-weight:700;flex-shrink:0;margin-top:3px;'
            'background:#8b949e14;color:var(--muted);border:1px solid var(--border)}\n'
            '.step-body{flex:1;min-width:0}\n'
            '.step-technique{font-size:.68rem;font-weight:600;text-transform:uppercase;'
            'letter-spacing:.5px;margin-bottom:3px;color:var(--muted)}\n'
            '.step-msg{font-size:.95rem;line-height:1.5;padding:7px 11px;background:var(--bg);'
            'border:1px solid var(--border);border-radius:10px;word-wrap:break-word}\n'
            '.step-note{background:#e3b34112;border-left:3px solid var(--yellow);padding:4px 10px;'
            'border-radius:0 6px 6px 0;margin:4px 0 2px;font-size:.76rem;color:var(--yellow)}\n'
            '.cp{background:none;border:1px solid transparent;color:var(--border);width:24px;height:24px;'
            'border-radius:6px;cursor:pointer;font-size:.7rem;flex-shrink:0;display:flex;align-items:center;'
            'justify-content:center;transition:.15s;position:relative;margin-top:7px;opacity:.3}\n'
            '.step:hover .cp{opacity:1;border-color:var(--border)}\n'
            '.cp:hover{background:var(--card);color:var(--text);opacity:1}\n'
            '.cp.ok{border-color:var(--green);color:var(--green);opacity:1}\n'
            '.cp.ok::after{content:"\\2713";position:absolute;font-size:.68rem;font-weight:700}\n'
            '.cp.ok .cp-icon{visibility:hidden}\n'
            '.top-btn{position:fixed;bottom:16px;right:16px;width:36px;height:36px;border-radius:50%;'
            'background:var(--card);border:1px solid var(--border);color:var(--muted);font-size:1rem;'
            'cursor:pointer;display:none;align-items:center;justify-content:center;z-index:100;transition:.15s}\n'
            '.top-btn:hover{color:var(--text);border-color:var(--muted)}.top-btn.show{display:flex}\n'
            '@media(max-width:700px){.page{padding:12px 10px}.hero{flex-direction:column;text-align:center}'
            '.toc{margin:0 -10px;padding-left:10px;padding-right:10px}.info-bar{justify-content:center}}\n'
            '</style>')

        # \u2500\u2500 JS \u2500\u2500
        JS = (
            '<script>\n'
            "document.addEventListener('click',e=>{const b=e.target.closest('.cp');if(!b)return;"
            "navigator.clipboard.writeText(b.dataset.copy).then(()=>{b.classList.add('ok');"
            "setTimeout(()=>b.classList.remove('ok'),1200)})});\n"
            "document.querySelectorAll('.protocol-header').forEach(h=>{"
            "h.addEventListener('click',()=>h.closest('.protocol').classList.toggle('collapsed'))});\n"
            "document.querySelectorAll('.seq-tab').forEach(tab=>{"
            "tab.addEventListener('click',()=>{"
            "const body=tab.closest('.protocol-body');"
            "body.querySelectorAll('.seq-tab').forEach(t=>t.classList.remove('active'));"
            "body.querySelectorAll('.seq-panel').forEach(p=>p.classList.remove('active'));"
            "tab.classList.add('active');"
            "document.getElementById(tab.dataset.seq).classList.add('active')})});\n"
            "document.querySelectorAll('.toc a').forEach(a=>{a.addEventListener('click',()=>{"
            "const t=document.getElementById(a.getAttribute('href').slice(1));"
            "if(t)t.classList.remove('collapsed')})});\n"
            "const topBtn=document.getElementById('top-btn');\n"
            "window.addEventListener('scroll',()=>topBtn.classList.toggle('show',window.scrollY>400));\n"
            "topBtn.addEventListener('click',()=>window.scrollTo({top:0,behavior:'smooth'}));\n"
            "const toc=document.getElementById('toc');\n"
            "window.addEventListener('scroll',()=>toc.classList.toggle('shadow',window.scrollY>150));\n"
            "const tocLinks=document.querySelectorAll('.toc a');\n"
            "const sections=Array.from(tocLinks).map(a=>({link:a,"
            "el:document.getElementById(a.getAttribute('href').slice(1))})).filter(s=>s.el);\n"
            "function updateToc(){let a=sections[0];for(const s of sections)"
            "if(s.el.getBoundingClientRect().top<=100)a=s;"
            "tocLinks.forEach(l=>l.classList.remove('active'));if(a)a.link.classList.add('active')}\n"
            "window.addEventListener('scroll',updateToc);updateToc();\n"
            '</script>')

        # \u2500\u2500 Assemble \u2500\u2500
        page = (
            '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
            '<meta charset="UTF-8">\n'
            '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
            '<title>%s \u2014 Objection & Resistance Handling</title>\n' % h(name_up)
            + CSS + '\n</head>\n<body>\n<div class="page">\n\n'
            '<a href="./" class="back-btn">\u2190 %s Guide</a>\n\n' % h(name)
            + '<div class="hero">\n'
            '  <img src="%s" alt="%s" class="hero-photo">\n'
            '  <div>\n    <h1>%s \u2014 Objection & Resistance</h1>\n'
            '    <div class="tagline">Complete sequences for Infloww</div>\n'
            '  </div>\n</div>\n\n'
            % (photo, h(name), h(name_up))
            + '<div class="info-bar">\n'
            '  <span class="info-pill"><b>%d</b> sheets</span>\n'
            '  <span class="info-pill"><b>%d</b> scripts</span>\n'
            '  <span class="info-pill"><b>%d</b> protocols</span>\n'
            '</div>\n\n' % (total_sheets, total_scripts, total_protos)
            + '<nav class="toc" id="toc">\n  '
            + '\n  '.join(toc_items) + '\n</nav>\n\n'
            + '\n'.join(proto_parts) + '\n\n'
            + '<div style="text-align:center;padding:24px 0 10px;color:var(--muted);font-size:.73rem">'
            'Chatting Wizard \u00b7 Script Manager \u00b7 2026<br>'
            '%s \u2014 OBJ/RES/SIT</div>\n\n' % h(name_up)
            + '</div>\n<button class="top-btn" id="top-btn">\u2191</button>\n\n'
            + JS + '\n</body>\n</html>')

        path = os.path.join(self.web_dir, "objections.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(page)
        print("[HTML] Objections: %s" % path)
        return path

    def generate_all(self):
        """Generate everything for this model."""
        print(f"\n{'='*60}")
        print(f"GENERATING: {self.c['name']}")
        print(f"{'='*60}")
        
        self.ensure_dirs()
        self.download_profile_photo()
        xlsx_path = self.generate_merged_xlsx()
        guide_path = self.generate_guide_html()
        obj_path = self.generate_objections_html()
        
        print(f"[DONE] {self.c['name']} complete!")
        return xlsx_path


if __name__ == "__main__":
    print("ModelFactory ready. Import and use with model configs.")
    print("Example: from model_factory import ModelFactory")
